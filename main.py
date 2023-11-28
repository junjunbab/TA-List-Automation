import os

#설치 완료후 환경설정
# python 패키지로 JAVA_HOME 설정하기
os.environ["JAVA_HOME"] = "/opt/conda"

# 필요 패키지 추가
import time
import datetime
import pickle
import docx
from docx import Document
import re
import sys, io
import tempfile
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection, GradientFill, Color
from openpyxl.cell import MergedCell
from openpyxl import load_workbook
from docx.shared import Pt, Cm
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.enum.section import WD_SECTION

# 한글 토큰화 konlpy 모듈
# from konlpy.tag import Okt
from ckonlpy.tag import Twitter
# okt = Okt()
twitter = Twitter() # twitter가 okt보다 성능 높음

# from gensim.models import Word2Vec

# 문서작업 중 필요 모듈
import re
import glob
import warnings
# import gensim

# model checkpoint
import mygsmod as gs
from mygsmod import ModelCheckpoint

import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
import tempfile
from PIL import Image

import streamlit as st

from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import confusion_matrix

import torch.nn as nn
import torch
import torch.optim as optim
import torchsummary
from torchsummary import summary
from torch.nn import functional as F
from torch.utils.data import Dataset, DataLoader
from torchtext.vocab import build_vocab_from_iterator
from torch.nn.utils.rnn import pad_sequence

# 직접 저장해둔 모듈 호출
import ta_eda_modeling
from ta_eda_modeling import word_map, noun_map, cr_word_map, stopwords
import ta_auto
from ta_auto import clear_directory, main_grouping, save_to_excel, categorize_sheet, group_and_save_data, drop_first_row, insert_row_based_on_condition, worklist_type_transform, split_excel_sheets_to_files, copy_style, cdv_checklist, hex_checklist, afc_checklist,integrate_docx_files, create_handbook_hex_afc, create_handbook_cdv, afc_checklist_merge,hex_checklist_merge,  cdv_checklist_merge

# 한글깨짐 방지코드 
font_location = '/home/sagemaker-user/gsc/NanumGothic.ttf'
fm.fontManager.addfont(font_location)
font_name = fm.FontProperties(fname=font_location).get_name()
matplotlib.rc('font', family=font_name)
matplotlib.rc('axes', unicode_minus=False)

# 웹 페이지 기본 설정
# page title: 데이터 분석 및 모델링 대시보드
st.set_page_config(
    page_title="TA Worklist & Checklist Automation", # page 타이틀
    page_icon="🧊", # page 아이콘
    layout="wide", # wide, centered
    initial_sidebar_state="auto", # 사이드 바 초기 상태
    menu_items={
        'Get Help': 'https://streamlit.io',
        'Report a bug': None,
        'About': '2023 GS CDS Class',
    }
)

# 실습 소개 페이지 출력 함수
# 소개 페이지는 기본으로 제공됩니다.
def front_page():
    st.title('TA Worklist & Checklist Automation Tool')
    st.header('Tool의 기능 소개')
    st.write('1. TA Report에서 Worklist 작성 사항 자동 추출')
    st.write('2. 작성된 Worklist로부터 Checklist 자동 생성')
    st.write('3. Checklist로부터 TA Handbook 자동 생성')
    st.markdown(' 1. EDA 페이지 생성')
    st.markdown('''
        - 파일 업로드 (TA Report Word 파일 업로드)
        - 파일 형식 변경 (Word 파일 -> txt 파일 -> DataFrame)
        - 데이터 전처리 (문장 분리, 단어 mapping, 한글화, 빈행 제거, 토큰화, 불용어 제거)
        - 추가 전처리 for 머신러닝 모델(CountVectorizer 함수를 통한 벡터화)
        - 추가 전처리 for 딥러닝 모델(Customized Twitter 함수를 통한 토큰화)
    ''')
    st.markdown(' 2. Modeling 페이지 생성')
    st.markdown('''
        - 머신러닝 모델 사용을 위한 데이터 분할
        - 딥러닝 모델 사용을 위한 데이터 분할
        - 모델링 (하이퍼 파라미터 설정)
        - 모델링 결과 확인 (평가 측도, Confusion Matrix)
    ''')
    st.markdown(' 3. Model Using 페이지 생성')
    st.markdown('''
        - 입력 값 설정 (메뉴)
        - 추론 
    ''')    
    
#-----------------------------------------------------------------------------------------
# customized_twitter에 특정 명사 입력
nouns = noun_map.values()
twitter.add_dictionary(nouns, 'Noun')

# 불용어 사전에 추가
plus_stopwords = ['도', '는', '다', '의', '가', '이','은', '한', '에', '하', '고', '을','를', '인', '듯', '과', '와', '네',
             '들', '듯', '지', '임', '게', '기', '개', '개소', '년', '번'
]
for word in plus_stopwords:
    stopwords.append(word)

# 부식율 인식 위한 사전
cr_word_map.update({round(i*0.001, 2): '높음' for i in range(101, 1000)})
cr_word_map[0.0] = '낮음'
cr_word_map[0.00] = '낮음'
#-----------------------------------------------------------------------------------------------------    
warnings.filterwarnings('ignore')  # 경고 메시지 무시
pd.set_option('display.max_colwidth', 700)  # pandas에서 표시하는 최대 열 너비 설정
#-----------------------------------------------------------------------------------------------------
# GPU 사용
device = 'cuda' if torch.cuda.is_available() else 'cpu'

# 시드 설정
torch.manual_seed(777)
torch.cuda.manual_seed_all(777)
#----------------------------------------------------------------------------------------------------
# 1. file load 함수
# 2. 파일 확장자에 맞게 읽어서 df으로 리턴하는 함수
# 3. 성능 향상을 위해 캐싱 기능 이용
@st.cache_data
def load_file(file):
    
    # 확장자 분리
    ext = file.name.split('.')[-1]
    
    # 확장자 별 로드 함수 구분
    if ext == 'docx':
        return pd.read_docx(file)
    
# word -> txt 함수
def add_requirement_to_sentences(text):
    # Split the text based on the pattern "(number) "
    sentences = re.split(r'(\d+\.\s|\(\d+\)\s)', text)
    
    # If no "(number) " pattern is found, treat the entire text as one sentence
    if len(sentences) == 1:
        return text.strip() + ' 필요.'

    # The split function will capture the delimiters as well, so we'll have to merge them back into the sentences
    processed_sentences = []
    for i in range(1, len(sentences), 2):
        sentence = sentences[i] + sentences[i+1].strip()
        if sentence.endswith('.'):
            sentence = sentence[:-1]
        sentence += ' 필요.'
        processed_sentences.append(sentence)
    
    # Join the sentences back together
    processed_text = ' '.join(processed_sentences)
    
    return processed_text

# 문서를 처리하는 함수
def process_document_txt(input_filepath, output_directory):
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    doc = Document(input_filepath)

    capture_next_table = False
    previous_paragraph = None
    filepath = ""

    for element in doc.element.body:
        if element.tag.endswith('p'):
            paragraph = element.text
            if paragraph:
                if '1. 장치 기본 정보' in paragraph and previous_paragraph is not None:
                    device_info = previous_paragraph
                    filename = "".join(x for x in device_info if x.isalnum() or x in " _-").rstrip()
                    filename += ".txt"
                    filepath = os.path.join(output_directory, filename)
                    with open(filepath, 'w', encoding='utf-8') as txt_file:
                        txt_file.write(previous_paragraph + '\n')

                if '2. 개방검사 결과' in paragraph:
                    capture_next_table = True

                previous_paragraph = paragraph

        elif element.tag.endswith('tbl') and capture_next_table:
            table = None
            for tbl in doc.tables:
                if tbl._element == element:
                    table = tbl
                    break

            if table and filepath:  # Ensure filepath is not empty
                with open(filepath, 'a', encoding='utf-8') as txt_file:
                    for row in table.rows[1:]:  # Skip the header row
                        first_column_text = row.cells[0].text.strip()
                        second_column_text = row.cells[1].text.strip()
                        # Apply special processing if the first column contains '차기 TA Recommend'
                        if first_column_text == '차기 TA Recommend':
                            if second_column_text not in ('N/A', ''):
                                second_column_text = add_requirement_to_sentences(second_column_text)
                        # Write the second column text to the file regardless of the first column content
                        txt_file.write(second_column_text + '\n')
                capture_next_table = False

        elif capture_next_table and ('3. 참고 사진' in paragraph or '3. 사진' in paragraph):
            capture_next_table = False
            
# txt -> df 함수
def custom_sentence_splitter(text):
    # '숫자' + '.' + '공백' or '숫자' + '.' + '문자' 문자를 기준으로 우선적으로 분리
    primary_sentences = re.split(r'(?<=\d)\.\s|(?<=\d)\.(?=[a-zA-Z\uAC00-\uD7A3])', text)
    
    refined_sentences = []
    for sent in primary_sentences:
        # 추가 분리: '숫자' + ')' + '공백'
        if re.search(r'\d\)\s', sent):
            parts = re.split(r'(?<=\d\))\s', sent)
            refined_sentences.extend(parts)
        # 추가 분리: '한글 문자 뒤에 오는 마침표(.) + '공백'
        elif re.search(r'[\uAC00-\uD7A3]\.\s', sent):
            parts = re.split(r'(?<=[\uAC00-\uD7A3])\.\s', sent)
            refined_sentences.extend(s + '.' for s in parts if s)  # 마침표 추가
        # 추가 분리: '한글 문자 뒤에 공백 2칸 이상일 경우 분리'
        elif re.search(r'[\uAC00-\uD7A3]\s\s', sent):
            parts = re.split(r'(?<=[\uAC00-\uD7A3])\s\s', sent)
            refined_sentences.extend(s + '.' for s in parts if s)
        # 추가 분리: '('+'숫자'+')' 분리'
        elif re.search(r'(\(\d+\)\s)', sent):
            parts = re.split(r'(?<=(\(\d+\)\s)', sent)
            refined_sentences.extend(s + '.' for s in parts if s) 
        else:
            refined_sentences.append(sent)
            
    # 선행 혹은 후행 공백과 빈 문자열 제거
    refined_sentences = [sent.strip() for sent in refined_sentences if sent.strip()]
    return refined_sentences

def process_text_files(path):
    all_files = glob.glob(os.path.join(path, '*.txt'))
    filename_list = []
    sent_list = []

    for file_ in all_files:
        with open(file_, 'r', encoding='utf-8') as f:
            first_line = f.readline().strip()
            remaining_text = f.read().strip()

        sentences = custom_sentence_splitter(remaining_text)
        sentences = [re.sub(r'([a-zA-Z])([\uAC00-\uD7A3])', r'\1 \2', sent) for sent in sentences]

        for sent in sentences:
            filename_list.append(first_line)
            sent_list.append(sent)

    return pd.DataFrame({'filename': filename_list, 'sent_text': sent_list})

# '.' 기준 문장 추가 분리
def split_sentences(text):
    # Divide the text into sentences based on the period (.) following the Hangul characters
    sentences = re.split('(?<=[\uAC00-\uD7A3])\.', text)
    sentences = [sent.strip() for sent in sentences if sent]  # Select only non-space sentences and remove leading and trailing spaces
    
    # Insert a space between English and Korean characters
    sentences = [re.sub(r'([a-zA-Z])([\uAC00-\uD7A3])', r'\1 \2', sent) for sent in sentences]
    return sentences

# 괄호 안의 문자 처리
def brackets_clean(text):
            # 괄호 안의 숫자, 특수기호만 제거
            clean1 = re.sub(r'\(([\d\W_]*?)\)', '()', text)
    
            # 괄호와 문자간 띄어쓰기
            clean2 = re.sub(r'([^\s])(\()', r'\1 \2', clean1)
            clean3 = re.sub(r'(\))([^\s])', r'\1 \2', clean2)
    
            return clean3

# 단어 변환 1 : word_map에서 단어 길이가 긴 순으로 먼저 변환 실시, re-tubing은 리튜빙으로, tubing은 튜빙으로 인식되도록 함
def replace(match):
    return word_map[match.group(0)]

def apply_replacement1(text):
    # word_map의 키를 길이에 따라 내림차순으로 정렬합니다.
    sorted_keys = sorted(word_map.keys(), key=len, reverse=True)
    # lookbehind와 lookahead를 사용하여 단어의 일부만 매치되도록 패턴을 수정합니다.
    pattern = re.compile('|'.join('(?<!\w){}(?!\w)'.format(re.escape(k)) for k in sorted_keys),re.IGNORECASE)
    return pattern.sub(replace, text)

# Function to apply the replacement within the text
def apply_replacement2(text):
    # Pattern that matches the words to be replaced even if they are part of a larger word
    pattern = re.compile('|'.join(map(re.escape, word_map.keys())))
    return pattern.sub(replace, text)

# 괄호 안의 문자 처리
def brackets_clean(text):
            # 괄호 안의 숫자, 특수기호만 제거
            clean1 = re.sub(r'\(([\d\W_]*?)\)', '()', text)
    
            # 괄호와 문자간 띄어쓰기
            clean2 = re.sub(r'([^\s])(\()', r'\1 \2', clean1)
            clean3 = re.sub(r'(\))([^\s])', r'\1 \2', clean2)
    
            return clean3

# default 찾기 부분 단어 변환 1 : word_map에서 단어 길이가 긴 순으로 먼저 변환 실시, re-tubing은 리튜빙으로, tubing은 튜빙으로 인식되도록 함
def replace_default(match):
    return word_map.get(match.group(0), match.group(0))

def apply_replacement1_default(text):
    # word_map의 키를 길이에 따라 내림차순으로 정렬합니다.
    sorted_keys = sorted(word_map.keys(), key=len, reverse=True)
    # lookbehind와 lookahead를 사용하여 단어의 일부만 매치되도록 패턴을 수정합니다.
    pattern = re.compile('|'.join('(?<!\w){}(?!\w)'.format(re.escape(k)) for k in sorted_keys),re.IGNORECASE)
    return pattern.sub(replace_default, text)

# default 찾기 부분 단어 변환 2
def apply_replacement2_default(text):
    # Pattern that matches the words to be replaced even if they are part of a larger word
    pattern = re.compile('|'.join(map(re.escape, word_map.keys())))
    return pattern.sub(replace_default, text)

# 부식율 인식 처리 함수
def replace_with_words(text, word_map):
    # Define a regular expression pattern for the intended formats
    pattern = r'(\d+\.\d{2})\s*mm(?:/yr|/year)?|(\d+\.\d{2})\*mm(?:/yr|/year)?'

    def replace(match):
        # Extract the number and round it
        num = round(float(match.group(1)), 2)
        # Replace with corresponding word if exists
        return f'{word_map.get(num, num)} mm/yr'  # Default to original number if not in word map

    # Replace all occurrences in the text
    return re.sub(pattern, replace, text)

# 한글 추출 함수
def extract_korean(text):
    hangul = re.compile('[^ ㄱ-ㅣ 가-힣]')  
    result = hangul.sub('', text)
    
    return result

# 머신러닝 적용 위한 토큰화 함수
def tokenize(doc):
    # pos 메서드를 사용하여 토큰화 및 품사 태깅, 정규화 및 기본형 변환 수행
    return [word for word, tag in twitter.pos(doc, norm=True, stem=True)]

# 딥러닝 적용 위한 토큰화 함수
def tokenize1(doc):
    # pos 메서드를 사용하여 토큰화 및 품사 태깅, 정규화 및 기본형 변환 수행
    return [word for word, tag in twitter.pos(doc, norm=True, stem=False)]

# file uploader 
# session_state에 다음과 같은 3개 값을 저장하여 관리함
# 1. st.session_state['eda_state'] = {}
#  1.1 : st.session_state['eda_state']['current_file']  / st.session_state['eda_state']['current_data']
# 2. st.session_state['modeling_state'] = {}
# 3. st.session_state['using_state'] = {}
def file_uploader():
    # 파일 업로더 위젯 추가 (Word 문서 선택)
    file = st.file_uploader("Select file (Word document)", type=['docx'])
    
    if file is not None:
        with st.spinner('변환 작업 중...'):
            # 새 파일 업로드 시 기존 상태 초기화
            st.session_state['eda_state'] = {}
            st.session_state['modeling_state'] = {}
            st.session_state['using_state'] = {}

            # 업로드된 파일을 임시 저장
            temp_file_path = "temp_uploaded_file.docx"
            with open(temp_file_path, "wb") as f:
                f.write(file.getbuffer())

            # word -> txt
            output_directory = "Word to Text File"
            process_document_txt(temp_file_path, output_directory)  # 수정된 함수 이름

            # 세션 상태 업데이트
            st.session_state['eda_state']['current_file'] = file
            st.success(f"**:blue[{output_directory} 변환 완료!!]**")

            # txt -> df
            df = process_text_files(output_directory)  # 수정된 함수 이름
            st.session_state['eda_state']['current_data'] = df
            st.success("Text to DataFrame 변환 완료!!")

        # 새로 업로드된 파일 로드
        if 'current_data' in st.session_state['eda_state']:
            df = st.session_state['eda_state']['current_data']
            # '.' 기준 문장 추가 분리
            df['changed_sent_text'] = df['sent_text'].apply(lambda x: split_sentences(x))
            df = df.explode('changed_sent_text').reset_index(drop=True)
            st.session_state['eda_state']['current_data'] = df
            st.success("DataFrame 문장분리 완료!!")
            # df 출력
            st.dataframe(st.session_state['eda_state']['current_data'])

# 공통 전처리 tab 함수
def preprocess1():
    if 'preprocess_started' not in st.session_state['eda_state']:
        st.session_state['eda_state']['preprocess_started'] = False
    if 'save_clicked' not in st.session_state['eda_state']:
        st.session_state['eda_state']['save_clicked'] = False
        
    if st.button('데이터 전처리 시작'):
        st.session_state['eda_state']['preprocess_started'] = True
        
    # Initialize state if not already done
    if 'name' not in st.session_state:
        st.session_state['eda_state']['name'] = None
    
    # if 'preprocess_started' in st.session_state['eda_state'] and st.session_state['eda_state']['preprocess_started']:
    if st.session_state['eda_state']['preprocess_started']:
        with st.spinner('전처리 작업 중...'):
            # Ensure that current_data is available in the session state
            if 'current_data' in st.session_state['eda_state']:
                try:
                    df = st.session_state['eda_state']['current_data']
                    # 소문자로 변환
                    df['changed_sent_text']= df['changed_sent_text'].str.lower()
                    # 공백 제거
                    df['changed_sent_text'] = df['changed_sent_text'].str.strip()
                    # 괄호 안 문자 처리
                    df['changed_sent_text'] = df['changed_sent_text'].apply(brackets_clean)
                    st.success("문장 전처리(괄호 내 문자 처리, 소문자 변환, 공백제거) 완료!!")

                    # 단어 변환 1 : word_map에서 단어 길이가 긴 순으로 먼저 변환 실시, re-tubing은 리튜빙으로, tubing은 튜빙으로 인식되도록 함
                    df['changed_sent_text'] = df['changed_sent_text'].apply(apply_replacement1)
                    st.success("단어 변환1 완료!!")

                    # 단어변환2 : 단어변환1에서 인식되지 않은 단어, 예를들어 전체tube 와 같은 문구 처리
                    df['changed_sent_text'] = df['changed_sent_text'].apply(apply_replacement2)
                    st.success("단어 변환2 완료!!")

                    # 부식율 인식되도록 처리
                    # 'changed_sent_text' 열에 함수를 적용하여 cr_word_map에 따라 숫자를 단어로 바꿉니다.
                    df['changed_sent_text'] = df['changed_sent_text'].apply(lambda x: replace_with_words(x, cr_word_map))
                    st.success("부식율 변환 완료!!")

                    # 토큰화를 위해 한글 문자만 추출
                    df['changed_sent_text'] = df['changed_sent_text'].apply(extract_korean)
                    st.success("한글 문자 추출 완료!!")

                    # 빈행 제거
                    df = df[df['changed_sent_text'].str.strip() != '']
                    st.success("비어있는 행 제거 완료!!")

                    # 세션 저장
                    st.session_state['eda_state']['current_data'] = df
                    st.success("모든 전처리 완료!!")
                    # df 출력
                    st.dataframe(df)

                    st.divider()
                    
                    st.session_state['eda_state']['name'] = st.text_input("저장할 파일명을 입력하세요 ex) 전처리된 306 ta report:", st.session_state['eda_state']['name'])

                    st.markdown('''
                            ※ 주의사항 : 반드시 '데이터 저장' 버튼 누르기 전 '저장할 파일명'을 입력 후 'Enter'키를 눌러주세요.
                    ''')

                    if st.button('데이터 저장'):
                        st.session_state['eda_state']['save_clicked'] = True

                    # 파일 저장
                    # if 'save_clicked' in st.session_state['eda_state'] and st.session_state['eda_state']['save_clicked']:
                    # if st.session_state['eda_state']['save_clicked']:
                    if st.session_state['eda_state']['save_clicked'] and 'current_data' in st.session_state['eda_state']:
                        with st.spinner('데이터 저장 중...'):
                            # Ensure the directory exists
                            output_dir = "./전처리된 ta report by Streamlit"
                            if not os.path.exists(output_dir):
                                os.makedirs(output_dir)

                            # Create the full file path
                            file_path = os.path.join(output_dir, f"{st.session_state['eda_state']['name']}.csv")

                            # Save to Excel
                            df.to_csv(file_path, index=False)
                            st.success(f"파일이 {file_path}에 저장되었습니다!!")

                except Exception as e:
                    st.error("에러 발생 : ", e)
    else:
        st.error("전처리 할 데이터가 없습니다!!")

            
# 추가 전처리 for 머신러닝 모델 tab 출력 함수
def preprocess2():     
    # 파일 업로더
    # 최대 용량은 서버 설정에서 변경 가능
    # https://docs.streamlit.io/library/advanced-features/configuration#set-configuration-options
    uploaded_file = st.file_uploader("전처리된 csv 파일을 업로드 하세요.", type=['csv'], key='unique_preprocess2_uploader')
    if uploaded_file is not None:
        # 새 파일 업로드 시 기존 상태 초기화
        st.session_state['modeling_state'] = {}
        st.session_state['using_state'] = {}
        
        # 데이터프레임 생성
        df = pd.read_csv(uploaded_file)
        st.session_state['eda_state']['ml_data'] = df
        st.write(df)
        
    # 새로 업로드된 파일 로드
    if 'ml_data' in st.session_state['eda_state']:
        with st.spinner('전처리 중...'): 
            df = st.session_state['eda_state']['ml_data']
            # 토큰화, 불용어 제거
            df['changed_sent_text'] = df['changed_sent_text'].astype(str).apply(tokenize)
            df['changed_sent_text'] = df['changed_sent_text'].apply(lambda x: [item for item in x if item not in stopwords])
            # 데이터가 2글자 이하인 행 삭제
            df = df[df['changed_sent_text'].apply(len) > 2]
            df = df.reset_index(drop = True)
            # 빈행 제거
            df = df[df['changed_sent_text'].str.strip() != '']

            # 세션 상태 업데이트
            st.session_state['eda_state']['ml_data'] = df
            st.success("머신러닝을 위한 토큰화, 불용어제거, 2글자 이하의 행 제거, 비어있는 행 제거 완료!!")
            
            st.divider()
            
            # df 출력
            st.dataframe(st.session_state['eda_state']['ml_data'])
            
            st.divider()
            
            with st.spinner('데이터 분할 중...'):    
                # 머신러닝 적용을 위한 벡터화
                df = st.session_state['eda_state']['ml_data']
                cnt_vect = CountVectorizer()
                bow_vect = cnt_vect.fit_transform(df['changed_sent_text'].astype(str))
                word_list = cnt_vect.get_feature_names_out()
                count_list = bow_vect.toarray().sum(axis=0)
                tag_df = df.drop(['filename','changed_sent_text', 'sent_text'],axis=1,inplace=False)
                feature_df = bow_vect
            
                # train, test 분리
                # 원본 데이터 프레임의 인덱스를 저장합니다.
                indices = df.index
                x_train, x_test, y_train, y_test, idx_train, idx_test = train_test_split(
                    feature_df, 
                    tag_df, 
                    indices, 
                    test_size=0.3, 
                    random_state=0
                )

                # x_test에 해당하는 원본 데이터셋의 행 인덱스
                bow_vect_df = pd.DataFrame(bow_vect)
                cond = bow_vect_df.iloc[idx_test].index
                df_xtest = df[['filename','changed_sent_text']].iloc[cond]

                # 세션 상태 업데이트
                st.session_state['modeling_state'] = {
                    'x_train1': x_train,
                    'y_train1': y_train,
                    'x_test1': x_test,
                    'y_test1': y_test,
                    'df_xtest': df_xtest
                }
                st.success('분할 완료')
                
                st.divider()
                
                st.write(f"x_train: {x_train.shape}, y_train: {y_train.shape}, x_test: {x_test.shape}, y_test: {y_test.shape}")
                
                st.divider()
                
                st.dataframe(df_xtest)

# 모델 생성
class DNNModel(nn.Module):
    def __init__(self):
        super().__init__()
        
        # 신경망 레이어를 정의
        self.fc1 = nn.Linear(1000, 128)  # 첫 번째 히든 레이어
        self.fc2 = nn.Linear(128, 64)    # 두 번째 히든 레이어
        self.fc3 = nn.Linear(64, 32)    # 두 번째 히든 레이어
        self.fc4 = nn.Linear(32, 16)    # 두 번째 히든 레이어
        # self.fc5 = nn.Linear(32, 16)    # 세 번째 히든 레이어
        # self.fc5 = nn.Linear(32, 16)    # 세 번째 히든 레이어
        # self.fc6 = nn.Linear(32, 16)    # 세 번째 히든 레이어
        self.output = nn.Linear(16, 1)  # 출력 레이어
        
        # 활성화 함수를 정의
        self.relu = nn.ReLU()           # ReLU 활성화 함수
        self.sigmoid = nn.Sigmoid()     # 시그모이드 활성화 함수
        
        # 레이어의 가중치를 Xavier uniform 방식으로 초기화
        nn.init.xavier_uniform_(self.fc1.weight)
        nn.init.xavier_uniform_(self.fc2.weight)
        nn.init.xavier_uniform_(self.fc3.weight)
        nn.init.xavier_uniform_(self.fc4.weight)
        # nn.init.xavier_uniform_(self.fc5.weight)
        # nn.init.xavier_uniform_(self.fc6.weight)
        nn.init.xavier_uniform_(self.output.weight)
        
    def forward(self, x):
        # 순전파를 정의
        out = self.relu(self.fc1(x))    # 첫 번째 레이어를 통과한 뒤 ReLU 적용
        out = self.relu(self.fc2(out))  # 두 번째 레이어를 통과한 뒤 ReLU 적용
        out = self.relu(self.fc3(out)) 
        out = self.relu(self.fc4(out))
        # out = self.relu(self.fc5(out))
        # out = self.relu(self.fc6(out))
        out = self.sigmoid(self.output(out))  # 출력 레이어를 통과한 뒤 시그모이드 적용
        return out
    
# 딥러닝 CustomDataset 클래스 생성
class CustomDataset(Dataset):
    def __init__(self, x, y):
        super().__init__() # 부모 클래스의 생성자를 호출
        
        self.x = x # 독립 변수(입력 데이터)
        self.y = y # 종속 변수(레이블)
        
    def __len__(self):
        return len(self.x) # 데이터셋의 전체 길이를 반환
    
    def __getitem__(self, idx):
        return self.x[idx], self.y[idx] # 인덱스에 해당하는 데이터와 레이블을 반환  

# 사전 토큰화된 데이터에 대한 생성기 함수를 정의    
def tokens_generator(data):
    for tokens in data:
        yield tokens
        
# 추가 전처리 for 딥러닝 모델 tab 출력 함수
def preprocess3():     
    # 파일 업로더
    # 최대 용량은 서버 설정에서 변경 가능
    # https://docs.streamlit.io/library/advanced-features/configuration#set-configuration-options
    uploaded_file = st.file_uploader("전처리된 csv 파일을 업로드 하세요.", type=['csv'], key='unique_preprocess3_uploader')
    if uploaded_file is not None:
         # 새 파일 업로드 시 기존 상태 초기화
        st.session_state['modeling_state'] = {}
        st.session_state['using_state'] = {}
        
        # 데이터프레임 생성
        df = pd.read_csv(uploaded_file)
        st.session_state['eda_state']['dp_data'] = df
        st.write(df)

    # 새로 업로드된 파일 로드
    if 'dp_data' in st.session_state['eda_state']:
        with st.spinner('추가 전처리 중...'):
            df = st.session_state['eda_state']['dp_data']
            # 토큰화, 불용어 제거
            df['changed_sent_text'] = df['changed_sent_text'].astype(str).apply(tokenize1)
            df['changed_sent_text'] = df['changed_sent_text'].apply(lambda x: [item for item in x if item not in stopwords])
            # 데이터가 2글자 이하인 행 삭제
            df = df[df['changed_sent_text'].apply(len) > 2]
            df = df.reset_index(drop = True)
            # 빈행 제거
            df = df[df['changed_sent_text'].str.strip() != '']

            # 세션 상태 업데이트
            st.session_state['eda_state']['dp_data'] = df
            st.success("딥러닝을 위한 토큰화, 불용어제거, 2글자 이하의 행 제거, 비어있는 행 제거 완료!!")
            # df 출력
            st.dataframe(df)
    
            # train, test 분리
            # 원본 데이터 프레임의 인덱스를 저장합니다.
            with st.spinner('데이터 분할 중...'):
                train_size = int(df.shape[0]*0.8) # 전체 데이터의 80%를 학습 데이터 크기로 설정
                train = df.sample(len(df), random_state=0)[:train_size] # 학습 데이터를 무작위로 추출
                test = df.sample(len(df), random_state=0)[train_size:] # 테스트 데이터를 무작위로 추출
                st.session_state['eda_state'] = {
                    'train' : train,
                    'test' : test,
                }
                # 독립변수, 종속변수 분할 후 array로 변환
                x_train, y_train = train['changed_sent_text'].values, train['tag'].values
                x_test, y_test = test['changed_sent_text'].values, test['tag'].values

            # 세션 상태 업데이트
            st.session_state['modeling_state'] = {
                'x_train2': x_train,
                'y_train2': y_train,
                'x_test2': x_test,
                'y_test2': y_test,
            }

            st.success('분할 완료')
        st.write(f"train, test 형태 확인 ... x_train: {x_train.shape}, y_train: {y_train.shape}, x_test: {x_test.shape}, y_test: {y_test.shape}")
        
        st.divider()
        
        # tag value_counts() 시각화
        def plot_value_counts(df, title):
            plt.figure(figsize=(10, 6))
            sns.barplot(x=df.index, y=df.values)
            plt.title(title)
            plt.ylabel('Frequency')
            plt.xlabel('Tags')
            st.pyplot(plt)
            
        with st.expander('Train and Test 데이터셋 tag 빈도 확인', expanded=False):
            train_tag_counts = train['tag'].value_counts(normalize=True)
            test_tag_counts = test['tag'].value_counts(normalize=True)

            # Displaying the dataframes
            pd.set_option('display.max_colwidth', None)
            st.subheader("Train 데이터셋 tag 빈도")
            plot_value_counts(train_tag_counts, "Train Set")

            st.subheader("Test 데이터셋 tag 빈도")
            plot_value_counts(test_tag_counts, "Test Set")
        
        if ('x_train2' in st.session_state['modeling_state'] and
            'y_train2' in st.session_state['modeling_state'] and
            'x_test2' in st.session_state['modeling_state'] and
            'y_test2' in st.session_state['modeling_state']):
            
            x_train = st.session_state['modeling_state']['x_train2']
            y_train = st.session_state['modeling_state']['y_train2']
            x_test = st.session_state['modeling_state']['x_test2']
            y_test = st.session_state['modeling_state']['y_test2']
            
            # 학습 데이터와 검증 데이터를 CustomDataset 인스턴스로 변환 후 세션 상태 업데이트
            st.session_state['modeling_state']['train_set'] = CustomDataset(x_train, y_train)
            st.session_state['modeling_state']['test_set'] = CustomDataset(x_test, y_test)
            st.success('학습데이터, 검증데이터 CustomDataset 인스턴스로 변환 완료')
        
            # 이터레이터에서 어휘를 구축
            vocab = build_vocab_from_iterator(
                iterator=tokens_generator(x_train), # 훈련 데이터로부터 토큰의 이터레이터를 생성
                max_tokens=1000, # 최대 토큰 수를 지정
                specials=['<unk>', '<sos>', '<eos>', '<pad>'] # 특수 토큰을 지정
            )

            # 알 수 없는 토큰에 대한 기본 인덱스를 설정
            vocab.set_default_index(vocab['<unk>'])
        
            # 세션 상태 업데이트
            st.session_state['modeling_state']['vocab'] = vocab
            st.success('단어사전 생성 완료')
            st.write(f'단어사전 개수 : {len(vocab.get_stoi())}개')
            
# EDA 페이지 출력 함수
def eda_page():
    st.title('Exploratory Data Analysis')
    
    # eda page tab 설정
    t1, t2, t3, t4 = st.tabs(['파일 업로드 및 형식 변경', '데이터 전처리', '추가 전처리 for 머신러닝 모델', '추가 전처리 for 딥러닝 모델'])
    
    with t1:
        file_uploader()
    
    with t2:
        preprocess1()
    
    with t3:
        preprocess2()
        
    with t4:
        preprocess3()
    
# train_model
def train_model1(selected_model, model_name):
    if ('x_train1' in st.session_state['modeling_state'] and 
        'x_test1' in st.session_state['modeling_state'] and 
        'y_train1' in st.session_state['modeling_state'] and 
        'y_test1' in st.session_state['modeling_state']):

        x_train = st.session_state['modeling_state']['x_train1']
        x_test = st.session_state['modeling_state']['x_test1']
        y_train = st.session_state['modeling_state']['y_train1']
        y_test = st.session_state['modeling_state']['y_test1']
        
        with st.spinner('학습 중...'): 
            model = LogisticRegression(random_state = 0)
            model.fit(x_train, y_train)
        st.success('학습 완료')

        with st.spinner('예측 값 생성 중...'):
            train_pred = model.predict(x_train)
            test_pred = model.predict(x_test)
        st.success('예측 값 생성 완료')
        
        # 모델 저장 경로 설정
        models_dir = './streamlit models'
        if not os.path.exists(models_dir):
            os.makedirs(models_dir)

        file_name = datetime.datetime.now().strftime('%Y%m%d%H%M')

        # 모델 파일 저장
        with open(f'{models_dir}/model_{model_name.replace(" ", "_")}_{file_name}.dat', 'wb') as f:
            pickle.dump(model, f)

        return model, y_train, train_pred, y_test, test_pred
        
# 머신러닝 modeling tab 출력 함수
def modeling1():
    model_list = ['Select Model', 'Logistic Regression']
    model_dict = {'Logistic Regression': LogisticRegression}
    selected_model = ''
    
    # Initialize state if not already done
    if 'selected_model' not in st.session_state:
        st.session_state['eda_state']['selected_model'] = None
        
    selected_model = st.selectbox('학습에 사용할 모델을 선택하세요.', model_list, index=0)
    
    if selected_model in model_list[1:]:
        with st.spinner('학습 중...'): 
            st.session_state['modeling_state']['selected_model'] = selected_model

            result = train_model1(model_dict[selected_model], selected_model)
            if result is not None:
                model, y_train, train_pred, y_test, test_pred = result
                st.session_state['modeling_state']['model1'] = model
                st.session_state['modeling_state']['y_train1'] = y_train
                st.session_state['modeling_state']['y_test1'] = y_test
                st.session_state['modeling_state']['train_pred1'] = train_pred
                st.session_state['modeling_state']['test_pred1'] = test_pred
                st.success('학습 종료')
            else:
                st.error('Model training failed or returned unexpected results.')  
            
# ML 결과 tab 함수
def ml_results():
    with st.expander('Metrics', expanded=True):
        if 'y_train1' in st.session_state['modeling_state']:
            st.divider()
            st.caption('Train Results')
            c1, c2, c3 = st.columns(3)
            left, right = c1.columns(2)
            ac = accuracy_score(st.session_state['modeling_state']['y_train1'], st.session_state['modeling_state']['train_pred1'])
            left.write('**:blue[Accuracy]**')
            right.write(f'{ac: 10.5f}')

            left, right = c2.columns(2)
            recall = recall_score(st.session_state['modeling_state']['y_train1'], st.session_state['modeling_state']['train_pred1'])
            left.write('**:blue[Recall]**')
            right.write(f'{recall: 10.5f}')

            left, right = c3.columns(2)
            f1 = f1_score(st.session_state['modeling_state']['y_train1'], st.session_state['modeling_state']['train_pred1'])
            left.write('**:blue[F1]**')
            right.write(f'{f1: 10.5f}')
        if 'y_test1' in st.session_state['modeling_state']:
            st.divider()
            st.caption('Test Results')
            c1, c2, c3 = st.columns(3)
            left, right = c1.columns(2)
            ac = accuracy_score(st.session_state['modeling_state']['y_test1'], st.session_state['modeling_state']['test_pred1'])
            left.write('**:blue[Accuracy]**')
            right.write(f'{ac: 10.5f}')

            left, right = c2.columns(2)
            recall = recall_score(st.session_state['modeling_state']['y_test1'], st.session_state['modeling_state']['test_pred1'])
            left.write('**:blue[Recall]**')
            right.write(f'{recall: 10.5f}')

            left, right = c3.columns(2)
            f1 = f1_score(st.session_state['modeling_state']['y_test1'], st.session_state['modeling_state']['test_pred1'])
            left.write('**:blue[F1]**')
            right.write(f'{f1: 10.5f}')
        
    st.divider()
    
    with st.expander('Result Visualization', expanded=False):
        if 'y_train1' in st.session_state['modeling_state']:
            
            confu = confusion_matrix(y_true = st.session_state['modeling_state']['y_test1'], y_pred = st.session_state['modeling_state']['test_pred1'])

            plt.figure(figsize=(4, 3))
            plot = sns.heatmap(confu, annot=True, annot_kws={'size':15}, cmap='OrRd', fmt='.10g')
            plt.title('Confusion Matrix')
            plt.ylabel('Actual')
            plt.xlabel('Predicted')
            fig = plot.get_figure()
            st.pyplot(fig)
            # Clear the current plot to avoid overlap with future plots
            plt.clf()
            
    st.divider()
    
    # 잘못 예측한 값 확인
    with st.expander('Result of incorrect prediciton', expanded=False):
        if 'df_xtest' in st.session_state['modeling_state']:
            # df['changed_sent_text'] 중 x_test 확인 용도
            df_xtest = st.session_state['modeling_state']['df_xtest']
            df_xtest = df_xtest.reset_index(drop=True)
        if 'x_test1' in st.session_state['modeling_state']:   
            x_test = st.session_state['modeling_state']['x_test1']
            x_test = pd.DataFrame(x_test)
        if 'y_test1' in st.session_state['modeling_state']: 
            y_test = st.session_state['modeling_state']['y_test1']
            y_test = pd.DataFrame(y_test).reset_index(drop=True)
        if 'test_pred1' in st.session_state['modeling_state']:   
            y_pred = st.session_state['modeling_state']['test_pred1']
            y_pred = pd.DataFrame(y_pred)

            df_xtest['x_test'] = x_test
            df_xtest['y_test'] = y_test
            df_xtest['y_pred'] = y_pred

            falsePositive = df_xtest[['changed_sent_text','y_test','y_pred']].loc[(df_xtest['y_test'] == 0) & (df_xtest['y_pred'] == 1)]
            falseNegative = df_xtest[['changed_sent_text','y_test','y_pred']].loc[(df_xtest['y_test'] == 1) & (df_xtest['y_pred'] == 0)]

            # pd.set_option('display.max_colwidth', None)
            # df 출력
            st.dataframe(falsePositive)
            st.dataframe(falseNegative)

# 미리 학습된 BiLSTM 모델 가져오기
def load_model(filepath):
    """
    Load the model from the given filepath and return it.
    """
    checkpoint = torch.load(filepath, map_location=torch.device('cpu'))
    model = DNNModel()  # Initialize your model
    model.load_state_dict(checkpoint['model_state_dict'])
    model.eval()

    return model

# 가져온 모델을 실행
def model_upload():
    # Specify the path to your model
    model_filepath = './streamlit models/dnn_model.pt'

    if os.path.exists(model_filepath):
        if st.button('모델 생성'):
            # Load the model
            model = load_model(model_filepath)
            st.session_state['modeling_state']['dp_model'] = model
            st.success("모델 가져오기 완료")
            # Now you can use the model for predictions or further processing
    else:
        st.error("Model file not found. Please check the file path.")

# 모델 평가 함수
def evaluate_model(model, data_loader, loss_function, device, return_labels=False):
    model = st.session_state['modeling_state']['dp_model'].to(device)
    # # 손실함수, 옵티마이저
    loss = nn.BCELoss().to(device)
    optimizer = optim.Adam(model.parameters(), lr=0.001)
    model.eval()
    total_loss, total_accuracy, total_f1, total_recall = 0, 0, 0, 0
    total_samples = len(data_loader.dataset)
    true_labels_list, predicted_labels_list = [], []
    
    # 검증 파트
    with torch.no_grad():
        for x, y in data_loader:
            x = x.to(torch.float32).to(device)
            y = y.to(torch.float32).to(device)

            outputs = model(x)
            cost = loss(outputs, y.view(-1, 1))

            total_loss += cost.item()
            predictions = (outputs > 0.5).to(torch.float32)
            correct_preds = (predictions == y.view(-1, 1)).sum().item()

            total_accuracy += correct_preds
            total_f1 += f1_score(y.cpu(), predictions.cpu())
            total_recall += recall_score(y.cpu(), predictions.cpu())

            if return_labels:
                true_labels_list.extend(y.view(-1).tolist())
                predicted_labels_list.extend(predictions.view(-1).tolist())

    avg_loss = total_loss / len(data_loader)
    avg_accuracy = total_accuracy / total_samples
    avg_f1 = total_f1 / len(data_loader)
    avg_recall = total_recall / len(data_loader)

    if return_labels:
        return avg_loss, avg_accuracy, avg_f1, avg_recall, true_labels_list, predicted_labels_list
    else:
        return avg_loss, avg_accuracy, avg_f1, avg_recall   
    
# 벡터화, 패딩        
def sent2seq(token_list, vocab):
    # 주어진 토큰 리스트를 사용하여 단어 사전에 기반한 인덱스의 리스트로 변환합니다.
    # 각 토큰은 단어 사전에 해당하는 숫자 인덱스로 변환됩니다.
    seq = [vocab[token] for token in token_list]
    return seq

def vectorize_seq(sequences, dimension=1000):
    # 각 시퀀스에 대한 원-핫 인코딩 벡터를 만듭니다.
    # 결과는 (시퀀스 개수, 단어 사전의 크기) 형태의 0으로 채워진 배열입니다.
    results = np.zeros((len(sequences), dimension))
    
    # 각 시퀀스에 대해, 해당하는 인덱스의 위치에 1을 설정합니다.
    # 이는 해당 단어가 문장 내에 존재함을 표시합니다.
    for i, seq in enumerate(sequences):
        results[i, seq] = 1.
    return results


def collate_function(batch, vocab):
    label_list = []
    sentence_list = []
    
    for (token_list, label) in batch:
        # 토큰 리스트를 단어 사전의 인덱스로 변환한 뒤 텐서로 변환합니다.
        seq = torch.tensor(vectorize_seq([sent2seq(token_list, vocab)])[0])
        sentence_list.append(seq)
        label_list.append(label)
    
    # pad_sequence를 사용하여 모든 시퀀스를 동일한 길이로 패딩합니다.
    # 'batch_first=True'는 배치 크기가 반환된 텐서의 첫 번째 차원이 됨을 의미합니다.
    # '<pad>' 토큰에 해당하는 인덱스를 사용하여 패딩합니다.
    seq_list = pad_sequence(sentence_list, padding_value=vocab['<pad>'], batch_first=True)
    
    # 레이블 리스트를 텐서로 변환합니다.
    label_list = torch.tensor(label_list)
    
    return seq_list, label_list  

# 딥러닝 modeling tab 출력 함수
def modeling2():
    if st.button('최적 모델 생성'):
        if 'dp_model' in st.session_state['modeling_state']:
            # 최적의 성능값으로 저장된 w,b 값을 이용하기
            model = st.session_state['modeling_state']['dp_model'].to(device)
            # 손실함수, 옵티마이저
            loss = nn.BCELoss().to(device)
            optimizer = optim.Adam(model.parameters(), lr=0.001)
            # # 최적 모델 가져오기
            with st.spinner('평가 중...'): 
                if 'modeling_state' in st.session_state and all(key in st.session_state['modeling_state'] for key in ['train_set', 'test_set', 'vocab']):
                    train_set = st.session_state['modeling_state']['train_set']
                    test_set = st.session_state['modeling_state']['test_set']
                    vocab = st.session_state['modeling_state']['vocab']

                    # 데이터로더 생성
                    train_loader = DataLoader(dataset=train_set, batch_size=64, shuffle=True, drop_last=True, collate_fn=lambda batch: collate_function(batch, vocab))
                    test_loader = DataLoader(dataset=test_set, batch_size=64, collate_fn=lambda batch: collate_function(batch, vocab))
                    st.session_state['modeling_state']['train_loader'] = train_loader
                    st.session_state['modeling_state']['test_loader'] = test_loader
                    st.session_state['modeling_state']['vocab'] = vocab
                    st.success('DataLoader 생성 완료')
                    
                    # Evaluation
                    train_loss, train_accuracy, train_f1, train_recall = evaluate_model(model, train_loader, loss, device)
                    test_loss, test_accuracy, test_f1, test_recall, true_labels, predicted_labels = evaluate_model(model, test_loader, loss, device, return_labels=True)
                    st.success('평가 지표 생성 완료')
                    
                    # 세션 상태 업데이트
                    st.session_state['modeling_state']['train_loss'] = train_loss
                    st.session_state['modeling_state']['test_loss'] = test_loss
                    st.session_state['modeling_state']['train_accuracy'] = train_accuracy
                    st.session_state['modeling_state']['test_accuracy'] = test_accuracy
                    st.session_state['modeling_state']['train_f1'] = train_f1
                    st.session_state['modeling_state']['test_f1'] = test_f1
                    st.session_state['modeling_state']['train_recall'] = train_recall
                    st.session_state['modeling_state']['test_recall'] = test_recall
                    st.session_state['modeling_state']['true_labels'] = true_labels
                    st.session_state['modeling_state']['predicted_labels'] = predicted_labels
                    
                else:
                    st.error("Modeling state not found or incomplete in the session state.")
        else:
            st.error("Model not found in the session state.")
            
# DL 결과 tab 함수
def dp_results():
    with st.expander('Metrics', expanded=True):
        if ('train_loss' in st.session_state['modeling_state'] and
            'test_loss' in st.session_state['modeling_state'] and
            'train_accuracy' in st.session_state['modeling_state'] and
            'test_accuracy' in st.session_state['modeling_state'] and
            'train_f1' in st.session_state['modeling_state'] and
            'test_f1' in st.session_state['modeling_state'] and
            'train_recall' in st.session_state['modeling_state'] and
            'test_recall' in st.session_state['modeling_state']):
            
            train_loss = st.session_state['modeling_state']['train_loss']
            test_loss = st.session_state['modeling_state']['test_loss']
            train_accuracy = st.session_state['modeling_state']['train_accuracy']
            test_accuracy = st.session_state['modeling_state']['test_accuracy']
            train_f1 = st.session_state['modeling_state']['train_f1']
            test_f1 = st.session_state['modeling_state']['test_f1']
            train_recall = st.session_state['modeling_state']['train_recall']
            test_recall = st.session_state['modeling_state']['test_recall']
            
            # Display results
            st.divider()
            st.caption('Train Results')
            c1, c2, c3, c4 = st.columns(4)
            left, right = c1.columns(2)
            left.write('**:blue[Loss]**')
            right.write(f'{train_loss: 10.5f}')
            
            left, right = c2.columns(2)
            left.write('**:blue[Accuracy]**')
            right.write(f'{train_accuracy: 10.5f}')

            left, right = c3.columns(2)
            left.write('**:blue[Recall]**')
            right.write(f'{train_recall: 10.5f}')

            left, right = c4.columns(2)
            left.write('**:blue[F1]**')
            right.write(f'{train_f1: 10.5f}')
        
            st.divider()
            st.caption('Test Results')
            c1, c2, c3, c4 = st.columns(4)
            left, right = c1.columns(2)
            left.write('**:blue[Val Loss]**')
            right.write(f'{test_loss: 10.5f}')
            
            left, right = c2.columns(2)
            left.write('**:blue[Val Accuracy]**')
            right.write(f'{test_accuracy: 10.5f}')

            left, right = c3.columns(2)
            left.write('**:blue[Val Recall]**')
            right.write(f'{test_recall: 10.5f}')

            left, right = c4.columns(2)
            left.write('**:blue[Val F1]**')
            right.write(f'{test_f1: 10.5f}')
        
    st.divider()
    
    with st.expander('Result Visualization', expanded=False):
        if ('true_labels' in st.session_state['modeling_state'] and
            'predicted_labels' in st.session_state['modeling_state']):
            true_labels = st.session_state['modeling_state']['true_labels']
            predicted_labels = st.session_state['modeling_state']['predicted_labels']
            
             # Visualization
            conf_matrix = confusion_matrix(true_labels, predicted_labels)
            plt.figure(figsize=(4, 3))
            sns.heatmap(conf_matrix, annot=True, fmt='g', cmap='Blues')
            plt.title('Confusion Matrix')
            plt.xlabel('Predicted labels')
            plt.ylabel('True labels')
            st.pyplot(plt)
            # Clear the current plot to avoid overlap with future plots
            plt.clf()
            
    st.divider()
    
    with st.expander('Incorrect Predictions', expanded=False):
        if ('true_labels' in st.session_state['modeling_state'] and
            'predicted_labels' in st.session_state['modeling_state'] and
            'test' in st.session_state['eda_state']):
            
            true_labels = st.session_state['modeling_state']['true_labels']
            predicted_labels = st.session_state['modeling_state']['predicted_labels']
            test = st.session_state['eda_state']['test']
            test['pred_tag'] = predicted_labels

            # Initializing lists to store indices
            false_positives_indices = []
            false_negatives_indices = []

            # Iterating to categorize incorrect predictions
            for i, (true, pred) in enumerate(zip(true_labels, predicted_labels)):
                if true != pred:
                    if true == 0 and pred == 1:
                        false_positives_indices.append(i)
                    elif true == 1 and pred == 0:
                        false_negatives_indices.append(i)

            # Retrieving rows from the dataframe
            false_positives_df = test[['changed_sent_text','tag','pred_tag']].iloc[false_positives_indices]
            false_negatives_df = test[['changed_sent_text','tag','pred_tag']].iloc[false_negatives_indices]

            # Displaying the dataframes
            pd.set_option('display.max_colwidth', None)
            st.subheader("False Positives:")
            st.dataframe(false_positives_df)

            st.subheader("False Negatives:")
            st.dataframe(false_negatives_df) # Find indices where predictions were incorrect
            
    st.divider()
    
# 입력된 텍스트 전처리 함수
def sentiment_predict(model, device, vocab):
    key = 'text_input_key'
    if key not in st.session_state:
        st.session_state[key] = ''
        
    # Initialize 'prediction_result' if it's not already in the session state
    if 'prediction_result' not in st.session_state:
        st.session_state['prediction_result'] = None
        
    new_sentence = st.text_input('text를 입력하세요: ', key=key)
    if new_sentence != st.session_state[key]:
        st.session_state[key] = new_sentence
        st.session_state.prediction_result = None  # Reset prediction result
    if new_sentence:
        new_sentence = re.sub(r'\(([\d\W_]*?)\)', '()', new_sentence)
        new_sentence = re.sub(r'([^\s])(\()', r'\1 \2', new_sentence)
        new_sentence = re.sub(r'(\))([^\s])', r'\1 \2', new_sentence)
        new_sentence = new_sentence.lower()
        new_sentence = new_sentence.strip()
        new_sentence = apply_replacement1(new_sentence)
        new_sentence = apply_replacement2(new_sentence)
        new_sentence = replace_with_words(new_sentence, cr_word_map)
        new_sentence = extract_korean(new_sentence)
        new_sentence = tokenize1(new_sentence)
        new_sentence = [item for item in new_sentence if item not in stopwords]
        # print(new_sentence)
    
        # Convert the tokenized sentence into a format suitable for the model
        encoded = collate_function([(new_sentence, 0)], vocab)
        input_ids = encoded[0].to(device).to(torch.float32)
    
        # Get the model output
        model.eval()
        with torch.no_grad():
            h = model(input_ids)
            # print(h)

        # Output the result
        if h.item() > 0.5:
            st.session_state.prediction_result = "이 문장은 {:.2f}% 확률로 긍정입니다".format(h.item() * 100)
        else:
            st.session_state.prediction_result = "이 문장은 {:.2f}% 확률로 부정입니다".format((1 - h.item()) * 100)
                
    # Display the result if it's available
    if st.session_state.prediction_result:
        st.write(st.session_state.prediction_result)    

# Modeling 페이지 출력 함수
def modeling_page():
    st.title('ML & DL Modeling')
    
    # tabs를 추가하세요.
    t1, t2, t3, t4 = st.tabs(['ML Modeling', 'ML Results', 'DL Modeling', 'DL Results'])

    # file upload tab 구현
    with t1:
        modeling1()
    
    with t2:
        ml_results()
        
    with t3:
        model_upload()
        modeling2()
        # Example usage
        st.divider()
        st.write('최적 모델 test 하기')
        if ('dp_model' in st.session_state['modeling_state'] and
            'vocab' in st.session_state['modeling_state']):
            model = st.session_state['modeling_state']['dp_model']
            vocab = st.session_state['modeling_state']['vocab']
            sentiment_predict(model, device, vocab)
    
    with t4:
        dp_results()
######################################################################################################
# 새로운데이터 전처리 함수1
def new_data_preprocess_predict(text, model, device, vocab):     
    # 새로 업로드된 파일 로드
    with st.spinner('추가 전처리 및 예측 중...'):
        # 토큰화, 불용어 제거
        text = tokenize1(text)
        text = [item for item in text if item not in stopwords]
        # 데이터가 2글자 이하인 행 삭제
        if len(text) <= 2:
            return None      
        
        # 모델의 입력 형식으로 변환
        encoded = collate_function([(text, 0)], vocab)
        # 변환된 데이터를 디바이스(CPU 또는 GPU)에 할당합니다.
        input_ids = encoded[0].to(device).to(torch.float32)

        # 모델을 평가 모드로 설정합니다.
        model.eval()

        with torch.no_grad():
            # 모델을 사용하여 예측을 수행합니다.
            h = model(input_ids)
             
    # 예측 결과에 따라 긍정(1) 또는 부정(0)을 반환합니다.
    return 1 if h.item() > 0.5 else 0

# 데이터 시각화 함수
# tag value_counts() 시각화
def plot_value_counts(df, title):
    plt.figure(figsize=(10, 6))
    sns.barplot(x=df.index, y=df.values)
    plt.title(title)
    plt.ylabel('Frequency')
    plt.xlabel('Tags')
    st.pyplot(plt)
    
# sheet_name cleaning funciton
def clean_sheet_name(name):    
    for char in['[', ']',':', '*','?','/', '\\','(', ')']:
        name =name.replace(char, '')
    # Convert name to lowercase
    name = name.lower()

    # Allow spaces only between numbers and letters
    cleaned_name = ""
    for i in range(len(name)):
        if name[i].isalnum() or (name[i] == " " and i > 0 and name[i-1].isalnum() and i < len(name)-1 and name[i+1].isalnum()):
            cleaned_name += name[i]

    return cleaned_name

# using 페이지 출력 함수
def using_page():
    st.title('Worklist 작성 및 ML & DL Model 사용')
    
    # tabs를 추가하세요.
    t1, t2, t3 = st.tabs(['Worklist 작성', 'ML Model 사용', 'DL Model 사용'])
    
    # file upload tab 구현
    with t1:
        nested_tab1, nested_tab2, nested_tab2 = st.tabs(["Worklist 가공1", "Worklist 가공2", "Worklist 작성"])
        with nested_tab1:
            uploaded_file = st.file_uploader("Worklist 작성에 참고하실 과거 Worklist를 업로드 하세요.", type=["csv", "xlsx"], key = 'worklist_upload')
            if uploaded_file is not None:
                with st.spinner('데이터 가공중...'): 
                    df = pd.read_excel(uploaded_file)
                    st.session_state['eda_state']['worklist_df'] = df

                    # '목차' 열에서 NaN 값이 있는 행 제거
                    df = df.dropna(subset=['목차'])
                    # '목차', '작업내용', '요청팀' 열만 선택
                    df = df[['목차', '작  업  내  용', '요청팀']]
                    # '요청팀' 열 필터링 조건 설정 ('장치Reliability1팀'이거나 NaN)
                    cond = (df['요청팀'] == '장치Reliability1팀') | (df['요청팀'].isna())
                    # 조건에 따라 '요청팀' 열 필터링
                    df['요청팀'] = df['요청팀'][cond]
                    # '목차' 열을 문자열 타입으로 변환
                    df['목차'] = df['목차'].astype(str)
                    
                    # Initialize an in-memory bytes buffer
                    output_buffer = io.BytesIO()
                    
                    # 엑셀 파일 쓰기를 위한 pd.ExcelWriter 설정
                    with pd.ExcelWriter(output_buffer, engine = 'xlsxwriter') as writer:
                        start_index = None  # 그룹의 시작 인덱스 초기화
                        # DataFrame의 각 행에 대해 반복
                        for index, row in df.iterrows():
                            # 특정 조건을 충족하는지 확인 (여기서는 문자열이 '.0'으로 끝나는 경우)
                            if str(row['목차']).endswith('.0'):
                                if start_index is not None:
                                    # 이전 그룹을 엑셀 시트로 작성
                                    group_df = df.loc[start_index:index - 1]
                                    # 첫 번째 행의 데이터로부터 시트 이름 생성 (이름 청소 함수 필요)
                                    sheet_name = clean_sheet_name(group_df.iloc[0]['작  업  내  용'])[:31]
                                    # 엑셀 파일에 작성
                                    group_df.to_excel(writer, sheet_name=sheet_name, index=False)

                                # 새 그룹의 시작 인덱스 업데이트
                                start_index = index

                        # 마지막 그룹이 존재하는 경우 파일에 작성
                        if start_index is not None and start_index < len(df):
                            group_df = df.loc[start_index:]
                            # 시트 이름 생성 (이름 청소 함수 필요)
                            sheet_name = clean_sheet_name(group_df.iloc[0]['작  업  내  용'])[:31]
                            # 엑셀 파일에 마지막 그룹 작성
                            group_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            st.success("Worklist 가공1 완료.")
                        
                    # Seek to the beginning of the stream
                    output_buffer.seek(0) 
                    # Download button for the Excel file
                    st.download_button(label="Download Excel file",
                                       data=output_buffer,
                                       file_name="가공1완료된 Worklist.xlsx",
                                       mime="application/vnd.ms-excel")
                    
                    
                    
#                     def merge_excel_files(file_paths):
#                         # This dictionary will hold dataframes with sheet names as keys
#                         combined_data = {}

#                         for file in file_paths:
#                             # Load each Excel file
#                             xls = pd.ExcelFile(file)

#                             # Iterate through each sheet in the Excel file
#                             for sheet_name in xls.sheet_names:
#                                 # Read each sheet
#                                 df = pd.read_excel(xls, sheet_name)

#                                 # If the sheet name is already in the dictionary, append the new data
#                                 if sheet_name in combined_data:
#                                     combined_data[sheet_name] = combined_data[sheet_name].append(df, ignore_index=True)
#                                 else:
#                                     # If this is the first time we're seeing this sheet name, add it to the dictionary
#                                     combined_data[sheet_name] = df

#                         # Save the combined data to a new Excel file
#                         writer = pd.ExcelWriter('통합된 worklist.xlsx', engine='openpyxl')

#                         for sheet_name, data in combined_data.items():
#                             # Write each dataframe to a different sheet
#                             data.to_excel(writer, sheet_name=sheet_name, index=False)

#                         writer.save()

#                     # Example usage
#                     file_paths = ['./8_1. worklist 분리/23년 분리된 파일.xlsx', './8_1. worklist 분리/19년 분리된 파일.xlsx', './8_1. worklist 분리/15년 분리된 파일.xlsx']
#                     merge_excel_files(file_paths)
    
    # file upload tab 구현
    with t2:
        uploaded_file = st.file_uploader("전처리가 완료된 새로운 데이터를 업로드 하세요.", type=["csv", "xlsx"], key = 'ml_upload')
        if uploaded_file is not None:
            
            df_new = pd.read_csv(uploaded_file)
            st.session_state['eda_state']['ml_verification_data'] = df_new
            st.success("데이터프레임 생성.")
            st.write(df_new)
            
            st.divider()
            
            with st.expander('학습되지 않은 새로운 데이터셋 tag 빈도 확인', expanded=False):
                tag_counts = df_new['tag'].value_counts(normalize=True)

                # Displaying the dataframes
                pd.set_option('display.max_colwidth', None)
                st.subheader("학습되지 않은 새로운 데이터셋 tag 빈도")
                plot_value_counts(tag_counts, "New DataSet")
                
            st.divider()
            
            with st.spinner('학습 데이터 생성 중...'):   
                if 'ml_data' in st.session_state['eda_state']:
                    df = st.session_state['eda_state']['ml_data']

                    # 머신러닝 적용을 위한 벡터화
                    df = st.session_state['eda_state']['ml_data']
                    cnt_vect = CountVectorizer()
                    bow_vect1 = cnt_vect.fit_transform(df['changed_sent_text'].astype(str))
                    word_list = cnt_vect.get_feature_names_out()
                    count_list = bow_vect1.toarray().sum(axis=0)
                    x_train = bow_vect1
                    y_train = df.drop(['filename','changed_sent_text', 'sent_text'],axis=1,inplace=False)
                    
                    st.success("학습 데이터 생성 완료.")
                    
                    st.divider()
                    
                    with st.spinner('검증 데이터 생성 중...'):
                        if 'ml_verification_data' in st.session_state['eda_state']:
                            df_new = st.session_state['eda_state']['ml_verification_data']

                            df_new['changed_sent_text'] = df_new['changed_sent_text'].astype(str).apply(tokenize)
                            df_new['changed_sent_text'] = df_new['changed_sent_text'].apply(lambda x: [item for item in x if item not in stopwords])
                            df_new = df_new[df_new['changed_sent_text'].apply(len) > 2]
                            df_new = df_new.reset_index(drop = True)
                            empty_rows = df_new[df_new['changed_sent_text'].astype(str).apply(lambda x: x.strip() == '')]
                            empty_rows_index = empty_rows.index
                            df_new = df_new.drop(empty_rows_index)

                            # 검증 데이터 생성
                            bow_vect2 = cnt_vect.transform(df_new['changed_sent_text'].astype(str))

                            x_valid = bow_vect2
                            y_valid = df_new.drop(['filename','changed_sent_text', 'sent_text'],axis=1,inplace=False)

                            x_valid.shape, y_valid.shape
                            
                            st.success("토큰화, 불용어제거, 2글자 이하의 행 제거 완료!!")
                            st.success("검증 데이터 생성 완료.")
                            
                            st.divider()
                            
                            with st.spinner('모델 생성 중...'):
                                
                                # 학습하기
                                # fit in training set
                                lr = LogisticRegression(random_state = 0)
                                lr.fit(x_train, y_train)
                                # predict in test set
                                y_pred = lr.predict(x_valid)

                                st.success('긍정, 부정 예측 완료.')

                                st.divider()
                        
                                with st.expander('Result Evaluation', expanded=False):

                                    st.caption('Train Results')
                                    c1, c2, c3 = st.columns(3)
                                    left, right = c1.columns(2)
                                    left.write('**:blue[Accuracy]**')
                                    right.write(f'{accuracy_score(y_valid, y_pred): 10.5f}')

                                    left, right = c2.columns(2)
                                    left.write('**:blue[Recall]**')
                                    right.write(f'{recall_score(y_valid, y_pred): 10.5f}')

                                    left, right = c3.columns(2)
                                    left.write('**:blue[F1]**')
                                    right.write(f'{f1_score(y_valid, y_pred): 10.5f}')
                                    
                                    st.divider()
                        
                                with st.expander('Result Visualization', expanded=False):

                                    # Creating the confusion matrix
                                    cm = confusion_matrix(y_valid, y_pred)

                                    # Plotting the confusion matrix
                                    plt.figure(figsize=(8, 6))
                                    plot = sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', xticklabels=['Negative', 'Positive'], yticklabels=['Negative', 'Positive'])
                                    plt.xlabel('Predicted')
                                    plt.ylabel('Actual')
                                    plt.title('Confusion Matrix of Sentiment Prediction')
                                    fig = plot.get_figure()
                                    st.pyplot(fig)
                                    # Clear the current plot to avoid overlap with future plots
                                    plt.clf()

                                    st.divider()

                                # 잘못 예측한 값 확인
                                with st.expander('Result of incorrect prediciton', expanded=False):
                                    
                                    df_new['pred_tag'] = y_pred

                                    falsePositive = df_new[['changed_sent_text','tag','pred_tag']].loc[(df_new['tag'] == 0) & (df_new['pred_tag'] == 1)]
                                    falseNegative = df_new[['changed_sent_text','tag','pred_tag']].loc[(df_new['tag'] == 1) & (df_new['pred_tag'] == 0)]

                                    # pd.set_option('display.max_colwidth', None)
                                    # df 출력
                                    st.subheader('falsePositive')
                                    st.dataframe(falsePositive)
                                    st.subheader('falseNegative')
                                    st.dataframe(falseNegative)
                                    
                                    st.divider()
                                    
                                # filename기준 구룹화 및 pred_tag 기준 분류하여 시각화
                                # Step 1: Data Preparation
                                # Group by 'filename' and then sort within groups by 'pred_tag'
                                grouped_df = df_new.groupby('filename', group_keys=False).apply(lambda x: x.sort_values('pred_tag', ascending=False))
                                
                                # 부정어는 빨간색으로 표시
                                def style_row(row):
                                    if row['pred_tag'] == 0:
                                        return f"<tr style='font-weight: bold; color: red;'><td>{row['filename']}</td><td>{row['sent_text']}</td><td>{row['pred_tag']}</td></tr>"
                                    else:
                                        return f"<tr><td>{row['filename']}</td><td>{row['sent_text']}</td><td>{row['pred_tag']}</td></tr>"

                                # Step 2: Streamlit App
                                with st.expander('Assistance for Worklist Writing', expanded=False):

                                    # Using Session State to store the index of the current file group
                                    if 'index' not in st.session_state:
                                        st.session_state['index'] = 0

                                    # Getting unique filenames to iterate through
                                    unique_filenames = sorted(df_new['filename'].unique())

                                    # Display current group
                                    current_file = unique_filenames[st.session_state['index']]
                                    st.subheader(f"Data for {current_file}")
                                    # Convert the DataFrame to HTML for custom styling
                                    styled_html = "<table><tr><th>Filename</th><th>Sent Text</th><th>Pred Tag</th></tr>"
                                    for _, row in grouped_df[grouped_df['filename'] == current_file].iterrows():
                                        styled_html += style_row(row)
                                    styled_html += "</table>"
                                    st.markdown(styled_html, unsafe_allow_html=True)

                                    # Navigation
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        if st.button("Previous"):
                                            if st.session_state['index'] > 0:
                                                st.session_state['index'] -= 1

                                    with col2:
                                        if st.button("Next"):
                                            if st.session_state['index'] < len(unique_filenames) - 1:
                                                st.session_state['index']+= 1
                            
                        else:
                            st.error("검증 데이터 생성 실패.")
                else:
                    st.error("학습 데이터 생성 실패.")
        else:
            st.error("업로드된 파일이 없습니다.")

    # file upload tab 구현
    with t3:
        uploaded_file = st.file_uploader("전처리가 완료된 새로운 데이터를 업로드 하세요.", type=["csv", "xlsx"], key = 'dp_upload')
        if uploaded_file is not None:
            
            df_new = pd.read_csv(uploaded_file)
            st.session_state['eda_state']['dp_verification_data'] = df_new
            st.success("데이터프레임 생성.")
            st.write(df_new)
            
            st.divider()
            
            with st.expander('학습되지 않은 새로운 데이터셋 tag 빈도 확인', expanded=False):
                tag_counts = df_new['tag'].value_counts(normalize=True)

                # Displaying the dataframes
                pd.set_option('display.max_colwidth', None)
                st.subheader("학습되지 않은 새로운 데이터셋 tag 빈도")
                plot_value_counts(tag_counts, "New DataSet")
                
            st.divider()
            
            with st.spinner('모델 생성 중...'):   
                if 'dp_model' in st.session_state['modeling_state']:
                    model = st.session_state['modeling_state']['dp_model'].to(device)
                    # 손실함수, 옵티마이저
                    loss = nn.BCELoss().to(device)
                    optimizer = optim.Adam(model.parameters(), lr=0.001)
                    st.success("Model 불러오기 성공.")
                    
                    st.divider()
            
                    if 'vocab' in st.session_state['modeling_state']:
                        vocab = st.session_state['modeling_state']['vocab']
                        df_new['pred_tag'] = df_new['changed_sent_text'].apply(lambda x: new_data_preprocess_predict(x, model, device, vocab) if pd.notnull(x) else None)
                        st.success("토큰화, 불용어제거, 2글자 이하의 행 제거 완료!!")
                        st.success("데이터 형태 변환 완료.")
                        st.success('긍정, 부정 예측 완료.')
                        
                        st.divider()
                        
                        with st.expander('Result Evaluation', expanded=False):
                            
                            true_labels = df_new['tag']
                            predicted_labels = df_new['pred_tag']

                            # Calculate metrics
                            accuracy = accuracy_score(true_labels, predicted_labels)
                            recall = recall_score(true_labels, predicted_labels, pos_label=1)
                            f1 = f1_score(true_labels, predicted_labels, pos_label=1)
                            
                            st.caption('Train Results')
                            c1, c2, c3 = st.columns(3)
                            left, right = c1.columns(2)
                            left.write('**:blue[Accuracy]**')
                            right.write(f'{accuracy: 10.5f}')

                            left, right = c2.columns(2)
                            left.write('**:blue[Recall]**')
                            right.write(f'{recall: 10.5f}')

                            left, right = c3.columns(2)
                            left.write('**:blue[F1]**')
                            right.write(f'{f1: 10.5f}')
                        
                        st.divider()
                        
                        with st.expander('Result Visualization', expanded=False):
                            
                            # Ensure there are no null values in the columns used for the confusion matrix
                            df_new = df_new.dropna(subset=['tag', 'pred_tag'])

                            # Creating the confusion matrix
                            cm = confusion_matrix(df_new['tag'], df_new['pred_tag'])

                            # Plotting the confusion matrix
                            plt.figure(figsize=(8, 6))
                            plot = sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', xticklabels=['Negative', 'Positive'], yticklabels=['Negative', 'Positive'])
                            plt.xlabel('Predicted')
                            plt.ylabel('Actual')
                            plt.title('Confusion Matrix of Sentiment Prediction')
                            fig = plot.get_figure()
                            st.pyplot(fig)
                            # Clear the current plot to avoid overlap with future plots
                            plt.clf()

                        st.divider()

                        # 잘못 예측한 값 확인
                        with st.expander('Result of incorrect prediciton', expanded=False):
                            
                            falsePositive = df_new[['changed_sent_text','tag','pred_tag']].loc[(df_new['tag'] == 0) & (df_new['pred_tag'] == 1)]
                            falseNegative = df_new[['changed_sent_text','tag','pred_tag']].loc[(df_new['tag'] == 1) & (df_new['pred_tag'] == 0)]

                            # pd.set_option('display.max_colwidth', None)
                            # df 출력
                            st.subheader('falsePositive')
                            st.dataframe(falsePositive)
                            st.subheader('falseNegative')
                            st.dataframe(falseNegative)
                            
                        st.divider()
                            
                        # filename기준 구룹화 및 pred_tag 기준 분류하여 시각화
                        # Step 1: Data Preparation
                        # Group by 'filename' and then sort within groups by 'pred_tag'
                        grouped_df = df_new.groupby('filename', group_keys=False).apply(lambda x: x.sort_values('pred_tag', ascending=False))

                        # 부정어는 빨간색으로 표시
                        def style_row(row):
                            if row['pred_tag'] == 0:
                                return f"<tr style='font-weight: bold; color: red;'><td>{row['filename']}</td><td>{row['sent_text']}</td><td>{row['pred_tag']}</td></tr>"
                            else:
                                return f"<tr><td>{row['filename']}</td><td>{row['sent_text']}</td><td>{row['pred_tag']}</td></tr>"

                        # Step 2: Streamlit App
                        with st.expander('Assistance for Worklist Writing', expanded=False):

                            # Using Session State to store the index of the current file group
                            if 'index' not in st.session_state:
                                st.session_state['index'] = 0

                            # Getting unique filenames to iterate through
                            unique_filenames = sorted(df_new['filename'].unique())

                            # Display current group
                            current_file = unique_filenames[st.session_state['index']]
                            st.subheader(f"Data for {current_file}")
                            # Convert the DataFrame to HTML for custom styling
                            styled_html = "<table><tr><th>Filename</th><th>Sent Text</th><th>Pred Tag</th></tr>"
                            for _, row in grouped_df[grouped_df['filename'] == current_file].iterrows():
                                styled_html += style_row(row)
                            styled_html += "</table>"
                            st.markdown(styled_html, unsafe_allow_html=True)

                            # Navigation
                            col1, col2 = st.columns(2)
                            with col1:
                                if st.button("Previous"):
                                    if st.session_state['index'] > 0:
                                        st.session_state['index'] -= 1

                            with col2:
                                if st.button("Next"):
                                    if st.session_state['index'] < len(unique_filenames) - 1:
                                        st.session_state['index']+= 1

                    else:
                        st.error("단어사전 불러오기 실패.")

                else:
                    st.error("Model 불러오기 실패.")
######################################################################################################
# Checklist 페이지 출력 함수
def checklist_page():
    st.title('Checklist Generator')
    
    if 'checklist_state' not in st.session_state:
        st.session_state['checklist_state'] = {}
        
    # Initialize state if not already done
    if 'grouped_file_path' not in st.session_state:
        st.session_state['checklist_state']['grouped_file_path'] = None
        
    # Initialize filename1 if not already done
    if 'filename1' not in st.session_state:
        st.session_state['checklist_state']['filename1'] = None
    
    # tabs를 추가하세요.
    t1, t2, t3, t4 = st.tabs(['Worklist 시트 별 분리', 'Worklist 형태 변경', 'Worklist 그루핑', 'Checklist 자동 생성'])

    # file upload tab 구현
    with t1:
        uploaded_file = st.file_uploader("최종 통합 Worklist의 Excel or CSV 파일을 업로드 해주세요.", type=["xlsx", "csv"])
        if uploaded_file is not None and st.button("Process File", key="worklist_sep_button1"):
             # 새 파일 업로드 시 기존 상태 초기화
            st.session_state['checklist_state'] = {}
            
            with st.spinner('분리 작업 중...'):
                processed_files = split_excel_sheets_to_files(uploaded_file)
                st.write('Worklist가 Sheet별로 분리되었습니다!!')
                # Save processed files in session state
                st.session_state['checklist_state']['processed_files'] = processed_files
                st.session_state['checklist_state']['selected_files'] = []
                
        if 'processed_files' in st.session_state['checklist_state'] and st.session_state['checklist_state']['processed_files']:
            # File selection
            file_names = [file_name for file_name, _ in st.session_state['checklist_state']['processed_files']]
            selected_files = st.multiselect('다운로드할 파일을 선택하세요:', file_names, key='file_selector')

            # Update selected files in session state
            st.session_state['checklist_state']['selected_files'] = selected_files

            # Download button
            if st.button('Download Selected Files', key="download_button1"):
                for file_name, file_data in st.session_state['checklist_state']['processed_files']:
                    if file_name in st.session_state['checklist_state']['selected_files']:
                        st.download_button(label=f"Download {file_name}", data=file_data, file_name=file_name, mime="application/vnd.ms-excel")
                        st.success('파일이 정상적으로 다운로드 되었습니다!!')
               
    with t2:
        # Place the file processing code inside this block
        uploaded_file = st.file_uploader("분리된 Worklist의 Excel or CSV 파일을 업로드 해주세요.", type=["csv", "xlsx"])
        file_name_input = st.text_input("다운로드 할 파일명을 정해주세요. ex) 변경된 62 No.3 CDU Worklist.xlsx", "processed_file.xlsx")

        if st.button('Process File', key="worklist_trans_button1") and uploaded_file is not None:
            with st.spinner('변경 작업 중...'):
            
                # Perform the initial transformation
                temp_file_path = worklist_type_transform(uploaded_file)

                # Define a path for the file after inserting rows
                _, intermediate_file_path = tempfile.mkstemp(suffix='.xlsx')

                # Insert rows based on condition
                insert_row_based_on_condition(temp_file_path, intermediate_file_path)

                # Define a path for the final file
                _, final_file_path = tempfile.mkstemp(suffix='.xlsx')

                # Drop the first row in the final file
                drop_first_row(intermediate_file_path, final_file_path)
                
                st.write('Worklist의 형태가 분리되었습니다!!')

                # Provide download link for the final file
                with open(final_file_path, 'rb') as file:
                    st.download_button(label="Download Files",
                                       data=file,
                                       file_name=file_name_input,
                                       mime="application/vnd.ms-excel")
                    st.write('파일이 정상적으로 다운로드 되었습니다!!')

                # Clean up temporary files
                os.remove(temp_file_path)
                os.remove(intermediate_file_path)
                os.remove(final_file_path)
        
    with t3:
        # 파일 업로더
        uploaded_file = st.file_uploader("변경된 Worklist의 Excel or CSV 파일을 업로드 해주세요.", type=["csv", "xlsx"])
        if st.button('Process File 1') and uploaded_file is not None:
            with st.spinner('그루핑 1 작업 중...'):
                
                # 파일 형식 결정
                file_type = 'csv' if uploaded_file.name.endswith('.csv') else 'xlsx'
                # group_and_save_data 함수 실행
                grouped_file_path = group_and_save_data(uploaded_file, file_type)
                # Save to session state
                st.session_state['checklist_state']['grouped_file_path'] = grouped_file_path 
                st.session_state['checklist_state']['prepared_file_paths'] = {}
                st.write('장치 별 Sheet 분리가 완료되었습니다!!')
        else:
                st.error('유효한 파일 경로가 없습니다. 파일을 다시 업로드하세요.')
        
        if st.session_state['checklist_state'].get('grouped_file_path'):
            with st.spinner('그루핑 2 작업 중...'):
                file_paths = main_grouping(st.session_state['checklist_state']['grouped_file_path'])

                # Store prepared files in session state
                for category, file_path in file_paths.items():
                    st.session_state['checklist_state']['prepared_file_paths'][category] = file_path

                # Clean up temporary files and session state
                del st.session_state['checklist_state']['grouped_file_path']
        else:
                st.error('유효한 파일 경로가 없습니다. 파일을 다시 업로드하세요.')
                
         # Create download buttons for prepared files
        for category, file_path in st.session_state['checklist_state'].get('prepared_file_paths', {}).items():
            with open(file_path, "rb") as file:
                st.download_button(label=f"Download {category} File", data=file, file_name=file_path, mime="application/vnd.ms-excel")
                
    with t4:
        # Create nested tabs within Main Tab 1
        nested_tab1, nested_tab2, nested_tab3 = st.tabs(["CDV Checklist", "HEX Checklist", 'AFC Checklist'])
        with nested_tab1:
            # 파일 업로드 위젯
            uploaded_file = st.file_uploader("그루핑된 CDV Worklist를 업로드 하세요.", type=["xlsx", "xls"])
            st.session_state['checklist_state']['filename1'] = st.text_input("다운로드 할 폴더명을 입력하세요. ex) 62공정 CDV:", st.session_state['checklist_state']['filename1'])
            
            # "Generate" 버튼
            if st.button("Generate Checklist", key="generate_checklist_button1"):
                if uploaded_file is not None:
                    
                    zip_file = cdv_checklist(uploaded_file, st.session_state['checklist_state']['filename1'])

                    # 생성된 zip 파일을 다운로드 링크로 제공
                    with open(zip_file, "rb") as f:
                        st.download_button(
                            label="Download Checklists",
                            data=f,
                            file_name="CDV Checklists.zip",
                            mime="application/zip"
                        )
                else:
                    st.error("Please upload a file.")
                    
        with nested_tab2:
            # 파일 업로드 위젯
            uploaded_file = st.file_uploader("그루핑된 HEX Worklist를 업로드 하세요.", type=["xlsx", "xls"])
            st.session_state['checklist_state']['filename1'] = st.text_input("다운로드 할 폴더명을 입력하세요. ex) 62공정 HEX:", st.session_state['checklist_state']['filename1'])
            
            # "Generate" 버튼
            if st.button("Generate Checklist", key="generate_checklist_button2"):
                if uploaded_file is not None:
                    
                    zip_file = hex_checklist(uploaded_file, st.session_state['checklist_state']['filename1'])

                    # 생성된 zip 파일을 다운로드 링크로 제공
                    with open(zip_file, "rb") as f:
                        st.download_button(
                            label="Download Checklists",
                            data=f,
                            file_name="HEX Checklists.zip",
                            mime="application/zip"
                        )
                else:
                    st.error("Please upload a file.")
                    
        with nested_tab3:
            # 파일 업로드 위젯
            uploaded_file = st.file_uploader("그루핑된 AFC Worklist를 업로드 하세요.", type=["xlsx", "xls"])
            st.session_state['checklist_state']['filename1'] = st.text_input("다운로드 할 폴더명을 입력하세요. ex) 62공정 AFC:", st.session_state['checklist_state']['filename1'])
            
            # "Generate" 버튼
            if st.button("Generate Checklist", key="generate_checklist_button3"):
                if uploaded_file is not None:
                    
                    zip_file = afc_checklist(uploaded_file, st.session_state['checklist_state']['filename1'])

                    # 생성된 zip 파일을 다운로드 링크로 제공
                    with open(zip_file, "rb") as f:
                        st.download_button(
                            label="Download Checklists",
                            data=f,
                            file_name="AFC Checklists.zip",
                            mime="application/zip"
                        )
                else:
                    st.error("Please upload a file.")

######################################################################################################
# handbook 페이지 출력 함수
def handbook_page():
    st.title('Handbook Generator')
    
    if 'handbook_state' not in st.session_state:
        st.session_state['handbook_state'] = {}
        
    # Initialize filename1 if not already done
    if 'title1' not in st.session_state:
        st.session_state['handbook_state']['title1'] = None

    # tabs를 추가하세요.
    t1, t2 = st.tabs(['Checklist 통합본 생성', 'Handbook 생성'])

    # file upload tab 구현
    with t1:
        nested_tab1, nested_tab2, nested_tab3 = st.tabs(["CDV Checklist 통합", "HEX Checklist 통합", 'AFC Checklist 통합'])
        with nested_tab1:
            uploaded_file = st.file_uploader("그루핑된 CDV Worklist의 Excel or CSV 파일을 업로드 해주세요.", type=["xlsx", "csv"])
            if uploaded_file and st.button("Process File", key="worklist_merge_button1"):
                with st.spinner('통합 작업 중...'):
                    saved_file_path = cdv_checklist_merge(uploaded_file)
                    st.write('Checklist 통합본이 생성되었습니다!!')

                    # Download button
                    with open(saved_file_path, "rb") as file:
                        btn = st.download_button(
                                label="Download Merged Checklist",
                                data=file,
                                file_name="CDV,HTR_Checklist_통합.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    if btn:
                        st.success("파일이 정상 다운로드 되었습니다.")
            
        with nested_tab2:
            uploaded_file = st.file_uploader("그루핑된 HEX Worklist의 Excel or CSV 파일을 업로드 해주세요.", type=["xlsx", "csv"])
            if uploaded_file and st.button("Process File", key="worklist_merge_button2"):
                with st.spinner('통합 작업 중...'):
                    saved_file_path = hex_checklist_merge(uploaded_file)
                    st.write('Checklist 통합본이 생성되었습니다!!')

                    # Download button
                    with open(saved_file_path, "rb") as file:
                        btn = st.download_button(
                                label="Download Merged Checklist",
                                data=file,
                                file_name="HEX_Checklist 통합.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    if btn:
                        st.success("파일이 정상 다운로드 되었습니다.")
                        
        with nested_tab3:
            uploaded_file = st.file_uploader("그루핑된 AFC Worklist의 Excel or CSV 파일을 업로드 해주세요.", type=["xlsx", "csv"])
            if uploaded_file and st.button("Process File", key="worklist_merge_button3"):
                with st.spinner('통합 작업 중...'):
                    saved_file_path = afc_checklist_merge(uploaded_file)
                    st.write('Checklist 통합본이 생성되었습니다!!')

                    # Download button
                    with open(saved_file_path, "rb") as file:
                        btn = st.download_button(
                                label="Download Merged Checklist",
                                data=file,
                                file_name="AFC_Checklist 통합.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    if btn:
                        st.success("파일이 정상 다운로드 되었습니다.")
                        
    with t2:
        nested_tab1, nested_tab2, nested_tab3 = st.tabs(["CDV Handbook 생성", "HEX / AFC Handbook 생성", 'Handbook 통합'])
        with nested_tab1:
            uploaded_file = st.file_uploader("CDV Checklist 통합파일을 업로드 해주세요.", type=["xlsx", "csv"])
            st.session_state['handbook_state']['title1'] = st.text_input('파일의 title을 작성해주세요. ex) 62공정 Checklist:', st.session_state['handbook_state']['title1'])
            if st.button("Process File", key="handbook_generate_button1"):
                if uploaded_file is not None: 
                    with st.spinner('Handbook 생성 중...'):
                        saved_file_path = create_handbook_cdv(uploaded_file, st.session_state['handbook_state']['title1'])
                        st.write('CDV Handbook이 생성되었습니다!!')

                        # Download button
                        with open(saved_file_path, "rb") as file:
                            btn = st.download_button(
                                    label="Download CDV Handbook",
                                    data=file,
                                    file_name="handbook.docx",
                                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                                )
                        if btn:
                            st.success("파일이 정상 다운로드 되었습니다.")
                            
        with nested_tab2:
            uploaded_file = st.file_uploader("HEX 또는 AFC Checklist 통합파일을 업로드 해주세요.", type=["xlsx", "csv"])
            # st.session_state['handbook_state']['title1'] = st.text_input('파일의 title을 작성해주세요. ex) 62공정 Checklist:', st.session_state['handbook_state']['title1'])
            if st.button("Process File", key="handbook_generate_button2"):
                if uploaded_file is not None: 
                    with st.spinner('Handbook 생성 중...'):
                        saved_file_path = create_handbook_hex_afc(uploaded_file)
                        st.write('HEX / AFC Handbook이 생성되었습니다!!')

                        # Download button
                        with open(saved_file_path, "rb") as file:
                            btn = st.download_button(
                                    label="Download HEX/AFC Handbook",
                                    data=file,
                                    file_name="handbook.docx",
                                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                                )
                        if btn:
                            st.success("파일이 정상 다운로드 되었습니다.")
                            
        with nested_tab3:
            uploaded_files = st.file_uploader("통합 하실 handbook 파일을 업로드 해주세요.   ※ 주의사항 : 통합 시킬 파일을 순서대로 업로드 해주세요.", type=['docx'], accept_multiple_files=True)
            # st.session_state['handbook_state']['title1'] = st.text_input('파일의 title을 작성해주세요. ex) 62공정 Checklist:', st.session_state['handbook_state']['title1'])
            if st.button("Integrate File", key="handbook_integrate_button1"):
                if uploaded_files is not None: 
                    with st.spinner('Handbook 통합 중...'):
                        integrated_doc_path = integrate_docx_files(uploaded_files)
                        st.write('Handbook이 통합되었습니다!!')

                        # Download button
                        with open(integrated_doc_path, "rb") as file:
                            btn = st.download_button(
                                    label="Download TA Handbook",
                                    data=file,
                                    file_name="handbook_integrated.docx",
                                    mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                                )
                        if btn:
                            st.success("파일이 정상 다운로드 되었습니다.")
                
######################################################################################################
                
# session_state에 사전 sidebar_state, eda_state, modeling_state, using_state를 추가하세요.
if 'sidebar_state' not in st.session_state:
    st.session_state['sidebar_state'] = {}
    st.session_state['sidebar_state']['current_page'] = front_page
if 'eda_state' not in st.session_state:
    st.session_state['eda_state'] = {}
if 'modeling_state' not in st.session_state:
    st.session_state['modeling_state'] = {}
if 'using_state' not in st.session_state:
    st.session_state['using_state'] = {}
if 'checklist_state' not in st.session_state:
    st.session_state['checklist_state'] = {}
if 'handbook_state' not in st.session_state:
    st.session_state['handbook_state'] = {}
    
# sidebar 추가
with st.sidebar:
    image = Image.open("./참조/칼텍스 로고.png")  # Replace with the path to your image
    st.image(image, use_column_width=True)
    st.subheader('Dashboard Menu')
    b1 = st.button('Front Page', use_container_width=True)
    b2 = st.button('EDA Page', use_container_width=True)
    b3 = st.button('Modeling Page', use_container_width=True)
    b4 = st.button('Using Page', use_container_width=True)
    st.divider()
    b5 = st.button('Checklist Page', use_container_width=True)
    b6 = st.button('Handbook Page', use_container_width=True)
    
if b1:
    st.session_state['sidebar_state']['current_page'] = front_page
#     st.session_state['sidebar_state']['current_page']()
    front_page()
elif b2:
    st.session_state['sidebar_state']['current_page'] = eda_page
#     st.session_state['sidebar_state']['current_page']()
    eda_page()
elif b3:
    st.session_state['sidebar_state']['current_page'] = modeling_page
#     st.session_state['sidebar_state']['current_page']()
    modeling_page()
elif b4:
    st.session_state['sidebar_state']['current_page'] = using_page
#     st.session_state['sidebar_state']['current_page']()
    using_page()
elif b5:
    st.session_state['sidebar_state']['current_page'] = checklist_page
#     st.session_state['sidebar_state']['current_page']()
    checklist_page()
elif b6:
    st.session_state['sidebar_state']['current_page'] = handbook_page
#     st.session_state['sidebar_state']['current_page']()
    handbook_page() 
else:
    st.session_state['sidebar_state']['current_page']()
