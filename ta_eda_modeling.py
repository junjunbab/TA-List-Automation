
import os

#설치 완료후 환경설정
# python 패키지로 JAVA_HOME 설정하기
os.environ["JAVA_HOME"] = "/opt/conda"

# 필요 패키지 추가
import time
import datetime
import pickle
import docx
from docx import Document
import re
import sys, io
import tempfile
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, Protection, GradientFill, Color
from openpyxl.cell import MergedCell
from openpyxl import load_workbook
from docx.shared import Pt, Cm
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.enum.section import WD_SECTION

# 한글 토큰화 konlpy 모듈
# from konlpy.tag import Okt
from ckonlpy.tag import Twitter
# okt = Okt()
twitter = Twitter() # twitter가 okt보다 성능 높음

# from gensim.models import Word2Vec

# 문서작업 중 필요 모듈
import re
import glob
import warnings
# import gensim

# model checkpoint
import mygsmod as gs
from mygsmod import ModelCheckpoint

import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
import tempfile

import streamlit as st

from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import confusion_matrix

import torch.nn as nn
import torch
import torch.optim as optim
import torchsummary
from torchsummary import summary
from torch.nn import functional as F
from torch.utils.data import Dataset, DataLoader
from torchtext.vocab import build_vocab_from_iterator
from torch.nn.utils.rnn import pad_sequence

# word_map, noun_map 저장
# 영어 문자 -> 한글 문자 변환
word_map = {
    'front': '앞', 'side': '사이드', 'tube': '튜브', 'sheet': '시트', 'sludge': '슬러지', 'plugging': '플러깅', 'rear': '후면', 'pitting': '공식',
    'corrosion': '부식', 'ferrel': '페룰', 'hydro': '수압', 'test': '시험', 'iris': '아이리스', 'shell': '쉘', 'demister': '데미스터',
    'pad': '패드', 'general': '일반', 'cover': '커버', 'channel': '찬넬', 'backing': '백킹', 'floating': '유동', 'head': '헤드', 'bundle': '번들',
    'baffle': '배플', 'tie': '타이', 'rod': '로드', 'spacer': '받침대', 'hard': '하드', 'scale': '스케일', 'gasket': '가스켓', 'face': '면',
    'partition': '파티션', 'groove': '그루브', 'top': '상부', 'bottom': '하부', 'tray': '트레이', 'power': '파워', 'brushing': '브러싱', 'stainless': '스테인리스',
    'foil': '호일', 'duct': '덕트', 'guide': '가이드', 'ring': '링', 'anchor': '앵커', 'fin': '핀', 'ceramic': '세라믹', 'wool': '울', 'flat': '플랫',
    'bar': '바', 'angle': '앵글', 'leak': '리크', 'expansion': '익스펜션', 'bellows': '벨로즈', 'flange': '플랜지', 'common': '커먼', 'rt': '알티',
    'seal': '씰', 'welding': '용접', 're tubing': '리튜빙', 'retubing': '리튜빙', 're-tubing': '리튜빙', 'roll': '롤', 'joint': '조인트', 'bed': '베드', 'distributor': '디스트리뷰터', 'packing': '팩킹',
    'support': '지지대', 'grid': '그리드', 'upper': '어퍼', 'lower': '로어', 'nozzle': '노즐', 'both': '양쪽', 'take out': '꺼냄', 'dent': '덴트',
    'depth': '깊이', 'touch': '터치', 'switching': '스위칭', 'insert': '삽입', 'duplex': '듀플렉스', 'spray': '스프레이', 'steam': '스팀', 'coil': '코일',
    'end': '엔드', 'plate': '판', 'ut': '유티', 'scanning': '스캐닝', 'errosion': '침식', 'plug': '플러그', 'hole': '홀', 'size': '사이즈',
    'coalescer': '코어레서', 'edge': '엣지', 'inside': '내부', 'pt': '피티', 'longitudinal': '길이방향', 'strip': '스트립', 'gap': '갭', 'vibration': '진동',
    'crack': '결함', 'bolting': '볼팅', 'sample': '샘플', 'desalter': '탈염기', 'performance': '퍼포먼스', 'neck': '넥', 'retubing': '리튜빙', 'filler': '필러',
    'film': '필름', 'wood': '우드', 'pipe': '파이프', 'branch': '브랜치', 'beam': '빔', 'sprinkler': '스프링쿨러', 'concrete': '콘크리트', 'wall': '벽',
    'pump': '펌프', 'suction': '흡입부분', 'screen': '스크린', 'mesh': '메쉬', 'boot': '부츠', 'phenolic': '페놀릭', 'epoxy': '에폭시',
    'bolt': '볼트', 'astm': '에이에스티엠', 'full': '가득', 'tag': '태그', 'stencil': '스텐실', 'heat': '히트', 'spare': '스페어', 'damage': '손상',
    'fixed': '고정식', 'type': '타입', 'elbow': '엘보', 'cutting': '커팅', 'high flux': '하이플럭스', 'coating': '코팅', 'oil': '유분', 'sand': '모래',
    'vacuum': '진공', 'trap': '트랩', 'inspection': '검사', 'door': '도어', 'open': '개방', 'cupronickel': '큐프로니켈', 'boss': '보스', 'small': '스몰',
    'liquid': '액체', 'seam': '심', 'griding': '그라인딩', 'heater': '히터', 'burner': '버너', 'radiant': '래디언트', 'hip': '힙', 'section': '부분',
    'ash': '애쉬', 'patch': '패치', 'oulet': '아웃렛', 'line': '라인', 'mixer': '믹서', 'down comer': '다운커머', 'vapor horn': '베이퍼혼', 'tpa': '티피에이',
    'reflux': '리플럭스', 'internal': '내부', 'monel': '모넬', 'clad': '클래드', 'declad': '디클래드', 'carbon steel': '탄소강', 'chimney': '침니',
    'main': '메인', 'zone': '영역', 'cladding': '클래드', 'feed': '피드', 'inlet': '인렛', 'sleeve': '슬리브', 'weir': '위어', 'column': '컬럼',
    'inhibitor': '인히비터', 'conical': '코니컬', 'decladding': '디클래드', 'buttering': '버터', 'grinding': '연마', 'overlay': '오버레이', 'final': '마지막',
    'lining': '라이닝', 'condensate': '응축수', 'scratch': '결함', 'pig': '피깅', 'decoking': '디코킹', 'coke': '코크', 'piping': '파이프', 'tank': '탱크',
    'header': '헤더', 'blind': '블라인드', 'manhole': '맨홀', 'boiler': '보일러', 'middle': '중앙', 'service':'서비스', 'ta':'대정비',
    'mt':'엠티', 'c/s':'탄소강', 'cs':'탄소강', '요함':'필요', '요망':'필요', '내화물':'내화재', '구성품':'부품', '열화':'열화',
    '부식공':'공식', '내부부식':'내부 부식', '최소값':'','s/d':'셧다운', 'edc':'이디씨', 'sa213-tp304l':'스테인리스',
    'expand':'확장', 'strength':'강도', '304l':'스테인리스','ss':'스테인리스', 'n/a':'', 'm/w':'입구','r.t':'알티', 'm.t':'엠티', 'p.t':'피티',
    'bt':'비티', 'toluene':'톨루엔', 'reboiler':'리보일러', 'structure':'구조물', 'maintenance':'보수',
	'scope':'범위', 'box':'박스', 'outlet':'토출부', 'new':'새로운', 'center':'중심', 'a179':'탄소강', 'heating':'열처리', 
	'heat':'열', 'carbon':'탄소강', 'erosion':'침식', 'rail':'레일', 'l-seam':'엘심', 'c-seam':'씨심', 'd-joint':'디조인트', 'p.t':'피티',
	'solvent':'솔벤트', 'cleaning':'청소', 'clamp':'클램프', 'band':'밴드', 'ect':'이씨티', 'pit':'공식', 'chloride':'염소',
	'316l':'스테인리스', 'collar':'칼라', 'bonnet':'보닛', 'water':'워터', 'drain':'드레인', 'air':'에어', 'blowing':'블로잉', 'normalizing':'노멀라이징',
	'cage':'케이지', 'random':'랜덤', 'girth':'거스', 'recovery':'복원', 'sliding':'슬라이딩', 'attachments':'부품', 'weld':'용접',
	'paut':'피에이유티', 'a516':'탄소강', 'solid':'일체형', 'device':'장치', 'notch':'노치', 'cu-ni':'구리 니켈', 'stacking':'스택형',
	'take-out':'꺼내', 'vertical':'수직', 'horizontal':'수평', 'spool':'스풀', 'sealing':'씰링', 'regenerator':'리제너레이터', 
	'unplugging':'언플러깅', 'impingement':'임핀지먼트', 'sparger':'스파저', 'design':'설계', 'mist':'미스트', 'eliminator':'앨리미네이터',
	'a213':'스테인리스', 'fluted':'세로홈', 'btm':'하부', 'scraper':'스크레퍼', 'agitator':'아지테이터', 'nut':'너트', 'load':'하중', 
    'reinforcement':'강화', 'rib':'리브', 'level':'레벨', 'cone':'콘', 'shaft':'샤프트','supervisor':'슈퍼바이저', 'steady':'스테디', 'bearing':'베어링', 'bracket':'브라켓',
    'manual':'매뉴얼', 'point':'부분', 'torque':'토크', 'twisted':'뒤틀린', 'packinox':'패키녹스', 'lifting':'리프팅', 'lug':'러그', '용접부mt':'용접부 mt',
    'manometer':'마노미터', 'holding':'홀딩', 'time':'시간', 'bubble':'버블', 'part':'부분', 'pass':'패스', 'chemical':'케미컬', 'tubesheet':'튜브시트',
    'convection':'컨벡션', 'casing':'케이싱', 'u-bend':'유벤드', 'chamber':'챔버', 'intermediate':'중간', 'casting':'캐스팅', 'stack':'스택', 'replica':'조직검사',
    'skin':'스킨', 'smooth':'스무스', 'mtpx':'엠티피엑스', 'reactor':'리액터', 'nox':'녹스', 'unlb':'유엔엘비', 'fuel':'연료', 'pilot':'파일럿', 'gas':'가스',
    'tip':'팁', 'tile':'타일', 'mid':'중간', 'sub':'보조', 'cooler':'쿨러', 'u-tube':'유튜브', 'low':'낮음', 'hammering':'해머링', 're-bolting':'재볼팅', 'fouling':'파울링',
    'hiflux':'하이플럭스', 'bare':'베어', 'high':'높음', 'flow':'플로우', 'scan':'스캔', 'snuffing':'스누핑', 'louver':'루버', 'galvanic':'갈바닉', 'panel':'판넬',
    'downcomer':'다운커머', 'surface':'표면', 'rung':'렁', 'contact':'접촉', 'stripper':'스트리퍼', 'receiver':'리시버', 'belzona':'벨조나', 'primary':'프라이머리',
    'secondary':'세컨더리', 'piece':'조각', 'double':'더블', 'inner':'내부', 'return':'리턴', 'bend':'벤드', 'item':'아이템', 'block':'블락', 'cold':'콜드', 'hot':'핫', 'dew':'이슬',
    'cement':'시멘트', 'fan':'팬', 'centrifuge':'센트리퓨지', 'shot':'쇼트',
    'silplate':'실플레이트', 'mounting':'마운팅', 'report':'레포트', 'corner':'코너', 'pinhole':'핀홀', 'metal':'메탈',
    'rough':'거친', 'brush':'브러쉬', 'staking':'스택형', 'handle':'핸들', 'access':'접근', 'over':'오버', 'shop':'샵','quench':'퀜치', 'go-pro':'고프로', 'pin':'핀', 
    'blistering':'블리스터링', 'brick':'내화재', 'valve':'밸브', 'cross':'크로스', 'sump':'섬프', 'unloading':'언로딩', 'electric':'전기', 'slot':'슬롯',
    'mixing':'혼합', 'cartridge':'카트리지', 'damper':'댐퍼', 'vortex':'볼텍스', 'tensioner':'텐셔너', 'take':'삽입', 
    'roof':'루프', 'nde':'비파괴검사', 'stationary':'고정식', 'attachment':'부품', 'hi-flex':'하이플렉스', 'hiflex':'하이플렉스', 'weep':'배수구', 'tp321':'스테인리스',
    'bending':'벤딩', 'adaptor':'어탭터', 'breaker':'브레이커', 'deflector':'디플렉터', 'scaning':'스캔', 'in-let':'인렛','out-let':'아웃렛','taper':'테이퍼', 'pigging':'피깅', 'loading':'로딩', 
    'blast':'블라스트', 'nortch':'노치', 'tap':'탭', 'connect':'커넥트', 'tofd':'티오에프디', 'dust':'먼지', 'pwht':'열처리', 'mass':'매스',
    'corosion':'부식', 'corrossion':'부식', 'corossion':'부식', 'graphite':'그라파이트', 'short':'짧은', 'long':'긴', 'term':'기간', 'thermowell':'써모웰',
    'vent':'벤트', 'castable':'단열', 'floor':'바닥', 'target':'타겟', 'thermocouple':'열전대', 'composite':'컴포짓',
    'data':'데이터', 'splitter':'스플리터', 'serration':'세레이션', 'observation':'관찰', 'flash':'플래쉬',
    'basket':'바스켓', 'manway':'입구', 'washer':'워셔', 'wraping':'래핑', 'injection':'인젝션', 'undercut':'언더컷',
    'mud':'머드', 'asphalt':'아스팔트', 'shield':'쉴드', 'static':'고정식', 'shoe':'슈', 'dump':'덤프', 'assembly':'부품',
    'cui':'보온재 부식', 'minor':'마이너', 'insulation':'보온재', 'beaker':'비커', 'cylinder':'실린더',
    'membrane':'멤브레인', 'hydraulic':'유압식', 'tension':'텐션', 'portable':'포터블', 'inconel':'인코넬',
    'ferrule':'페룰', 'machine':'머신', 'slag':'슬래그', 'plastic':'플라스틱', 'pressure':'압력',
    'bake':'베이크', 'out':'아웃', 'hold':'홀드', 'total':'토탈', 'saddle':'새들', 'release':'릴리즈',
    'torch':'토치', 'tesioning':'텐셔닝', 'washing':'워싱', 'embrittlement':'취성', 'skirt':'스커트',
    'grinder':'그라인더', 'repair':'보수', 'hopper':'호퍼', 'filter':'필터', 'force':'힘', 'glass':'글라스',
    'teflon':'테프론', 'knuckle':'너클', 'wrench':'렌치', 'soft':'소프트', 'east':'동쪽', 'west':'서쪽', 'south':'남쪽',
    'north':'북쪽', 'protection':'보호', 'hair':'헤어', 'clscc':'염화 부식', 'tight':'타이트', 'sponge':'스펀지',
    'tp304':'스테인리스', 'blushing':'브러싱', 'copper':'구리', 'location':'지역', 'vendor':'벤더',
    'recommendation':'검토 사항', 'cooling':'쿨링', 'gauging':'가우징', 'monitoring':'모니터링',
    'rust':'부식', 'grind':'그라인딩',
    'trouble':'트러블', 'breather':'브리더', 'critical':'중요한', 'v/v':'밸브', 'powder':'파우더', 
    'grating':'그레이팅', 'takeout':'분리', 'resin':'레진', 'bag':'백', 'rubber':'러버', 'follow up':'후속조치',
    'followup':'후속조치', 'follow-up':'후속조치', 'f/u':'후속조치', 'gauge':'게이지', 'ball':'볼',
    'blower':'블로워', 'wire':'와이어', 'jet':'제트', 'bortex':'볼텍스', 'cap':'캡', 'jacket':'자켓',
    'rotary':'로터리', 'blend':'블렌드', 'product':'생산품', 'extractor':'익스트렉터', 'flare':'플레어',
    'lance':'랜스', 'pp':'폴리프로필렌', 'belt':'벨트', 'pulling':'풀링', 'vane':'베인', 'propylene':'프로필렌',
    'sling':'슬링', 'upgrade':'업그레이드', 'tower':'타워', 'plenum':'플레넘', 'purge':'퍼지',
    'pellet':'펠릿', 'drilling':'드릴링', 'painting':'도색', 'sensor':'센서', 'effluent':'이플루언트', 'reducer':'레듀사', 'vapor':'증기',
    'close':'닫음', 'reject':'리젝트', 'scrapper':'스크레퍼',
    'stand':'스탠드', 'mash':'매쉬', 'imping':'임핀지먼트', 'tack':'택', 'preheat':'열처리', 'emissivity':'복사율',
    'detecting':'감지', 'economizer':'이코노마이저', 'manifold':'매니폴드',
    'combustible':'컴버스터블', 'catalyst':'촉매', 'thermiculate':'써미큘레이트', 'charging':'차징',
    'hastelloy':'하스텔로이', 'hanger':'행거', 'roughness':'거칠기', 'guillotine':'길로틴', 
    'thickness':'두께', 'spark':'스파크', 'chlorination':'염화 처리', 'superheater':'수퍼 히터', 'ejector':'이젝터',
    'regeneration':'리제너레이션', 'differential':'차이', 'reduction':'리덕션', 'drying':'드라잉', 'incoloy':'인콜로이', '파단':'파손'
}

noun_map = {
    'front': '앞', 'side': '사이드', 'tube': '튜브', 'sheet': '시트', 'sludge': '슬러지', 'plugging': '플러깅', 'rear': '후면', 'pitting': '공식',
    'corrosion': '부식', 'ferrel': '페룰', 'hydro': '수압', 'test': '시험', 'iris': '아이리스', 'shell': '쉘', 'demister': '데미스터',
    'pad': '패드', 'general': '일반', 'cover': '커버', 'channel': '찬넬', 'backing': '백킹', 'floating': '유동', 'head': '헤드', 'bundle': '번들',
    'baffle': '배플', 'tie': '타이', 'rod': '로드', 'spacer': '받침대', 'hard': '하드', 'scale': '스케일', 'gasket': '가스켓', 'face': '면',
    'partition': '파티션', 'groove': '그루브', 'top': '상부', 'bottom': '하부', 'tray': '트레이', 'power': '파워', 'brushing': '브러싱', 'stainless': '스테인리스',
    'foil': '호일', 'duct': '덕트', 'guide': '가이드', 'ring': '링', 'anchor': '앵커', 'fin': '핀', 'ceramic': '세라믹', 'wool': '울', 'flat': '플랫',
    'bar': '바', 'angle': '앵글', 'leak': '리크', 'expansion': '익스펜션', 'bellows': '벨로즈', 'flange': '플랜지', 'common': '커먼', 'rt': '알티',
    'seal': '씰', 'welding': '용접', 're tubing': '리튜빙', 'retubing': '리튜빙', 're-tubing': '리튜빙', 'roll': '롤', 'joint': '조인트', 'bed': '베드', 'distributor': '디스트리뷰터', 'packing': '팩킹',
    'support': '지지대', 'grid': '그리드', 'upper': '어퍼', 'lower': '로어', 'nozzle': '노즐', 'both': '양쪽', 'take out': '꺼냄', 'dent': '덴트',
    'depth': '깊이', 'touch': '터치', 'switching': '스위칭', 'insert': '삽입', 'duplex': '듀플렉스', 'spray': '스프레이', 'steam': '스팀', 'coil': '코일',
    'end': '엔드', 'plate': '판', 'ut': '유티', 'scanning': '스캐닝', 'errosion': '침식', 'plug': '플러그', 'hole': '홀', 'size': '사이즈',
    'coalescer': '코어레서', 'edge': '엣지', 'inside': '내부', 'pt': '피티', 'longitudinal': '길이방향', 'strip': '스트립', 'gap': '갭', 'vibration': '진동',
    'crack': '결함', 'bolting': '볼팅', 'sample': '샘플', 'desalter': '탈염기', 'performance': '퍼포먼스', 'neck': '넥', 'retubing': '리튜빙', 'filler': '필러',
    'film': '필름', 'wood': '우드', 'pipe': '파이프', 'branch': '브랜치', 'beam': '빔', 'sprinkler': '스프링쿨러', 'concrete': '콘크리트', 'wall': '벽',
    'pump': '펌프', 'suction': '흡입부분', 'screen': '스크린', 'mesh': '메쉬', 'boot': '부츠', 'phenolic': '페놀릭', 'epoxy': '에폭시',
    'bolt': '볼트', 'astm': '에이에스티엠', 'full': '가득', 'tag': '태그', 'stencil': '스텐실', 'heat': '히트', 'spare': '스페어', 'damage': '손상',
    'fixed': '고정식', 'type': '타입', 'elbow': '엘보', 'cutting': '커팅', 'high flux': '하이플럭스', 'coating': '코팅', 'oil': '유분', 'sand': '모래',
    'vacuum': '진공', 'trap': '트랩', 'inspection': '검사', 'door': '도어', 'open': '개방', 'cupronickel': '큐프로니켈', 'boss': '보스', 'small': '스몰',
    'liquid': '액체', 'seam': '심', 'griding': '그라인딩', 'heater': '히터', 'burner': '버너', 'radiant': '래디언트', 'hip': '힙', 'section': '부분',
    'ash': '애쉬', 'patch': '패치', 'oulet': '아웃렛', 'line': '라인', 'mixer': '믹서', 'down comer': '다운커머', 'vapor horn': '베이퍼혼', 'tpa': '티피에이',
    'reflux': '리플럭스', 'internal': '내부', 'monel': '모넬', 'clad': '클래드', 'declad': '디클래드', 'carbon steel': '탄소강', 'chimney': '침니',
    'main': '메인', 'zone': '영역', 'cladding': '클래드', 'feed': '피드', 'inlet': '인렛', 'sleeve': '슬리브', 'weir': '위어', 'column': '컬럼',
    'inhibitor': '인히비터', 'conical': '코니컬', 'decladding': '디클래드', 'buttering': '버터', 'grinding': '연마', 'overlay': '오버레이', 'final': '마지막',
    'lining': '라이닝', 'condensate': '응축수', 'scratch': '결함', 'pig': '피깅', 'decoking': '디코킹', 'coke': '코크', 'piping': '파이프', 'tank': '탱크',
    'header': '헤더', 'blind': '블라인드', 'manhole': '맨홀', 'boiler': '보일러', 'middle': '중앙', 'service':'서비스', 'ta':'대정비',
    'mt':'엠티', 'c/s':'탄소강', 'cs':'탄소강', '요함':'필요', '요망':'필요', '내화물':'내화재','구성품':'부품', '열화':'열화',
    '부식공':'공식', 's/d':'셧다운', 'edc':'이디씨', 'sa213-tp304l':'스테인리스','부':'부','shot':'쇼트',
    'expand':'확장', 'strength':'강도', '304l':'스테인리스', 'm/w':'입구', 'r.t':'알티', 'm.t':'엠티', 'p.t':'피티',
	'ss':'스테인리스', 'bt':'비티', 'toluene':'톨루엔', 'reboiler':'리보일러', 'structure':'구조물', 'maintenance':'보수',
	'scope':'범위', 'box':'박스', 'outlet':'토출부', 'new':'새로운', 'center':'중심', 'a179':'탄소강', 'heating':'열처리', 
	'heat':'열', 'carbon':'탄소강', 'erosion':'침식', 'rail':'레일', 'l-seam':'엘심', 'c-seam':'씨심', 'd-joint':'디조인트', 'p.t':'피티',
	'solvent':'솔벤트', 'cleaning':'청소', 'clamp':'클램프', 'band':'밴드', 'ect':'이씨티', 'pit':'공식', 'chloride':'염소',
	'316l':'스테인리스', 'collar':'칼라', 'bonnet':'보닛', 'water':'워터', 'drain':'드레인', 'air':'에어', 'blowing':'블로잉', 'normalizing':'노멀라이징',
	'cage':'케이지', 'random':'랜덤', 'girth':'거스', 'recovery':'복원', 'sliding':'슬라이딩', 'attachments':'부품', 'weld':'용접',
	'paut':'피에이유티', 'a516':'탄소강', 'solid':'일체형', 'device':'장치', 'notch':'노치', 'cu-ni':'구리 니켈', 'stacking':'스택형',
	'take-out':'꺼내', 'vertical':'수직', 'horizontal':'수평', 'spool':'스풀', 'sealing':'씰링', 'regenerator':'리제너레이터', 
	'unplugging':'언플러깅', 'impingement':'임핀지먼트', 'sparger':'스파저', 'design':'설계', 'mist':'미스트', 'eliminator':'앨리미네이터',
	'a213':'스테인리스', 'fluted':'세로홈', 'btm':'하부', 'scraper':'스크레퍼', 'agitator':'아지테이터', 'nut':'너트', 'load':'하중', 
    'reinforcement':'강화', 'rib':'리브', 'level':'레벨', 'cone':'콘', 'shaft':'샤프트','supervisor':'슈퍼바이저', 'steady':'스테디', 'bearing':'베어링', 'bracket':'브라켓',
    'manual':'매뉴얼', 'point':'부분', 'torque':'토크', 'twisted':'뒤틀린', 'packinox':'패키녹스', 'lifting':'리프팅', 'lug':'러그', '용접부mt':'용접부 mt',
    'manometer':'마노미터', 'holding':'홀딩', 'time':'시간', 'bubble':'버블', 'part':'부분', 'pass':'패스', 'chemical':'케미컬', 'tubesheet':'튜브시트',
    'convection':'컨벡션', 'casing':'케이싱', 'u-bend':'유벤드', 'chamber':'챔버', 'intermediate':'중간', 'casting':'캐스팅', 'stack':'스택', 'replica':'조직검사',
    'skin':'스킨', 'smooth':'스무스', 'mtpx':'엠티피엑스', 'reactor':'리액터', 'nox':'녹스', 'unlb':'유엔엘비', 'fuel':'연료', 'pilot':'파일럿', 'gas':'가스',
    'tip':'팁', 'tile':'타일', 'mid':'중간', 'sub':'보조', 'cooler':'쿨러', 'u-tube':'유튜브', 'low':'낮음', 'hammering':'해머링', 're-bolting':'재볼팅', 'fouling':'파울링',
    'hiflux':'하이플럭스', 'bare':'베어', 'high':'높음', 'flow':'플로우', 'scan':'스캔', 'snuffing':'스누핑', 'louver':'루버', 'galvanic':'갈바닉', 'panel':'판넬',
    'downcomer':'다운커머', 'surface':'표면', 'rung':'렁', 'contact':'접촉', 'stripper':'스트리퍼', 'receiver':'리시버', 'belzona':'벨조나', 'primary':'프라이머리',
    'secondary':'세컨더리', 'piece':'조각', 'double':'더블', 'inner':'내부', 'return':'리턴', 'bend':'벤드', 'item':'아이템', 'block':'블락', 'cold':'콜드', 'hot':'핫', 'dew':'이슬',
    'cement':'시멘트', 'fan':'팬', 'centrifuge':'센트리퓨지','silplate':'실플레이트', 'mounting':'마운팅', 'report':'레포트', 'corner':'코너', 'pinhole':'핀홀', 'metal':'메탈',
    'rough':'거친', 'brush':'브러쉬', 'staking':'스택형', 'handle':'핸들', 'access':'접근', 'over':'오버', 'shop':'샵','quench':'퀜치', 'go-pro':'고프로', 'pin':'핀', 
    'blistering':'블리스터링', 'brick':'내화재', 'valve':'밸브', 'cross':'크로스', 'sump':'섬프', 'unloading':'언로딩', 'electric':'전기', 'slot':'슬롯', 'in':'내부', 'out':'외부',
    'mixing':'혼합', 'cartridge':'카트리지', 'damper':'댐퍼', 'vortex':'볼텍스', 'tensioner':'텐셔너', 'take':'삽입', 
    'roof':'루프', 'nde':'비파괴검사', 'stationary':'고정식', 'attachment':'부품', 'hi-flex':'하이플렉스', 'hiflex':'하이플렉스', 'weep':'배수구', 'tp321':'스테인리스',
    'bending':'벤딩', 'adaptor':'어탭터', 'breaker':'브레이커', 'deflector':'디플렉터', 'scaning':'스캔', 'in-let':'인렛', 'out-let':'아웃렛', 'taper':'테이퍼', 'pigging':'피깅', 'loading':'로딩', 
    'blast':'블라스트', 'nortch':'노치', 'tap':'탭', 'connect':'커넥트', 'tofd':'티오에프디', 'dust':'먼지', 'pwht':'열처리', 'mass':'매스',
    'corosion':'부식', 'graphite':'그라파이트', 'short':'짧은', 'long':'긴', 'term':'기간', 'thermowell':'써모웰',
    'vent':'벤트', 'castable':'단열', 'floor':'바닥', 'target':'타겟', 'thermocouple':'열전대', 'composite':'컴포짓',
    'data':'데이터', 'splitter':'스플리터', 'serration':'세레이션', 'observation':'관찰', 'flash':'플래쉬',
    'basket':'바스켓', 'manway':'입구', 'washer':'워셔', 'wraping':'래핑', 'injection':'인젝션', 'undercut':'언더컷',
    'mud':'머드', 'asphalt':'아스팔트', 'shield':'쉴드', 'static':'고정식', 'shoe':'슈', 'dump':'덤프', 'assembly':'부품',
    'cui':'보온재 부식', 'minor':'마이너', 'insulation':'보온재', 'beaker':'비커', 'cylinder':'실린더',
    'membrane':'멤브레인', 'hydraulic':'유압식', 'tension':'텐션', 'portable':'포터블', 'inconel':'인코넬',
    'ferrule':'페룰', 'machine':'머신', 'slag':'슬래그', 'plastic':'플라스틱', 'pressure':'압력',
    'bake':'베이크', 'out':'아웃', 'hold':'홀드', 'total':'토탈', 'saddle':'새들', 'release':'릴리즈',
    'torch':'토치', 'tesioning':'텐셔닝', 'washing':'워싱', 'embrittlement':'취성', 'skirt':'스커트',
    'grinder':'그라인더', 'repair':'보수', 'hopper':'호퍼', 'filter':'필터', 'force':'힘', 'glass':'글라스',
    'teflon':'테프론', 'knuckle':'너클', 'wrench':'렌치', 'soft':'소프트', 'east':'동쪽', 'west':'서쪽', 'south':'남쪽',
    'north':'북쪽', 'protection':'보호', 'hair':'헤어', 'clscc':'염화 부식', 'tight':'타이트', 'sponge':'스펀지',
    'tp304':'스테인리스', 'blushing':'브러싱', 'copper':'구리', 'location':'지역', 'vendor':'벤더',
    'recommendation':'검토 사항', 'cooling':'쿨링', 'gauging':'가우징', 'monitoring':'모니터링',
    'rust':'부식', 'grind':'그라인딩', 'ejector':'이젝터',
    'trouble':'트러블', 'breather':'브리더', 'critical':'중요한', 'v/v':'밸브', 'powder':'파우더', 
    'grating':'그레이팅', 'takeout':'분리', 'resin':'레진', 'bag':'백', 'rubber':'러버', 'follow up':'후속조치',
    'followup':'후속조치', 'follow-up':'후속조치', 'f/u':'후속조치', 'gauge':'게이지', 'ball':'볼',
    'blower':'블로워', 'wire':'와이어', 'jet':'제트', 'bortex':'볼텍스', 'cap':'캡', 'jacket':'자켓',
    'rotary':'로터리', 'blend':'블렌드', 'product':'생산품', 'extractor':'익스트렉터', 'flare':'플레어',
    'lance':'랜스', 'pp':'폴리프로필렌', 'belt':'벨트', 'pulling':'풀링', 'vane':'베인', 'propylene':'프로필렌',
    'sling':'슬링', 'upgrade':'업그레이드', 'tower':'타워', 'plenum':'플레넘', 'purge':'퍼지',
    'pellet':'펠릿', 'drilling':'드릴링', 'painting':'도색', 'sensor':'센서', 'effluent':'이플루언트', 'reducer':'레듀사', 'vapor':'증기',
    'close':'닫음', 'reject':'리젝트', 'scrapper':'스크레퍼',
    'stand':'스탠드', 'mash':'매쉬', 'imping':'임핀지먼트', 'tack':'택', 'preheat':'열처리', 'emissivity':'복사율',
    'detecting':'감지', 'economizer':'이코노마이저', 'manifold':'매니폴드',
    'combustible':'컴버스터블', 'catalyst':'촉매', 'thermiculate':'써미큘레이트', 'charging':'차징',
    'hastelloy':'하스텔로이', 'hanger':'행거', 'roughness':'거칠기', 'guillotine':'길로틴', 
    'thickness':'두께', 'spark':'스파크', 'chlorination':'염화 처리', 'superheater':'수퍼 히터',
    'regeneration':'리제너레이션', 'differential':'차이', 'reduction':'리덕션', 'drying':'드라잉', 'incoloy':'인콜로이', '파단':'파손'
}

# cr_word_map, 불용어 사전 저장
# cr_word_map (부식율 처리)
# 부식율 인식을 위한 처리(0.1까지는 낮음으로 인식 되도록)
# 부식율 인식을 위한 처리(0.1까지는 낮음으로 인식 되도록)
cr_word_map = {round(i*0.001, 2): '낮음' for i in range(1, 101)}

# 불용어 사전
stopwords = pd.read_csv('./참조/불용어사전.csv').values.tolist()

# # word -> txt 함수
# def add_requirement_to_sentences(text):
#     # Split the text based on the pattern "(number) "
#     sentences = re.split(r'(\d+\.\s|\(\d+\)\s)', text)
    
#     # If no "(number) " pattern is found, treat the entire text as one sentence
#     if len(sentences) == 1:
#         return text.strip() + ' 필요.'

#     # The split function will capture the delimiters as well, so we'll have to merge them back into the sentences
#     processed_sentences = []
#     for i in range(1, len(sentences), 2):
#         sentence = sentences[i] + sentences[i+1].strip()
#         if sentence.endswith('.'):
#             sentence = sentence[:-1]
#         sentence += ' 필요.'
#         processed_sentences.append(sentence)
    
#     # Join the sentences back together
#     processed_text = ' '.join(processed_sentences)
    
#     return processed_text

# # 문서를 처리하는 함수
# def process_document_txt(input_filepath, output_directory):
#     if not os.path.exists(output_directory):
#         os.makedirs(output_directory)

#     doc = Document(input_filepath)

#     capture_next_table = False
#     previous_paragraph = None
#     filepath = ""

#     for element in doc.element.body:
#         if element.tag.endswith('p'):
#             paragraph = element.text
#             if paragraph:
#                 if '1. 장치 기본 정보' in paragraph and previous_paragraph is not None:
#                     device_info = previous_paragraph
#                     filename = "".join(x for x in device_info if x.isalnum() or x in " _-").rstrip()
#                     filename += ".txt"
#                     filepath = os.path.join(output_directory, filename)
#                     with open(filepath, 'w', encoding='utf-8') as txt_file:
#                         txt_file.write(previous_paragraph + '\n')

#                 if '2. 개방검사 결과' in paragraph:
#                     capture_next_table = True

#                 previous_paragraph = paragraph

#         elif element.tag.endswith('tbl') and capture_next_table:
#             table = None
#             for tbl in doc.tables:
#                 if tbl._element == element:
#                     table = tbl
#                     break

#             if table and filepath:  # Ensure filepath is not empty
#                 with open(filepath, 'a', encoding='utf-8') as txt_file:
#                     for row in table.rows[1:]:  # Skip the header row
#                         first_column_text = row.cells[0].text.strip()
#                         second_column_text = row.cells[1].text.strip()
#                         # Apply special processing if the first column contains '차기 TA Recommend'
#                         if first_column_text == '차기 TA Recommend':
#                             if second_column_text not in ('N/A', ''):
#                                 second_column_text = add_requirement_to_sentences(second_column_text)
#                         # Write the second column text to the file regardless of the first column content
#                         txt_file.write(second_column_text + '\n')
#                 capture_next_table = False

#         elif capture_next_table and ('3. 참고 사진' in paragraph or '3. 사진' in paragraph):
#             capture_next_table = False
            
# # txt -> df 함수
# def custom_sentence_splitter(text):
#     # '숫자' + '.' + '공백' or '숫자' + '.' + '문자' 문자를 기준으로 우선적으로 분리
#     primary_sentences = re.split(r'(?<=\d)\.\s|(?<=\d)\.(?=[a-zA-Z\uAC00-\uD7A3])', text)
    
#     refined_sentences = []
#     for sent in primary_sentences:
#         # 추가 분리: '숫자' + ')' + '공백'
#         if re.search(r'\d\)\s', sent):
#             parts = re.split(r'(?<=\d\))\s', sent)
#             refined_sentences.extend(parts)
#         # 추가 분리: '한글 문자 뒤에 오는 마침표(.) + '공백'
#         elif re.search(r'[\uAC00-\uD7A3]\.\s', sent):
#             parts = re.split(r'(?<=[\uAC00-\uD7A3])\.\s', sent)
#             refined_sentences.extend(s + '.' for s in parts if s)  # 마침표 추가
#         # 추가 분리: '한글 문자 뒤에 공백 2칸 이상일 경우 분리'
#         elif re.search(r'[\uAC00-\uD7A3]\s\s', sent):
#             parts = re.split(r'(?<=[\uAC00-\uD7A3])\s\s', sent)
#             refined_sentences.extend(s + '.' for s in parts if s)
#         # 추가 분리: '('+'숫자'+')' 분리'
#         elif re.search(r'(\(\d+\)\s)', sent):
#             parts = re.split(r'(?<=(\(\d+\)\s)', sent)
#             refined_sentences.extend(s + '.' for s in parts if s) 
#         else:
#             refined_sentences.append(sent)
            
#     # 선행 혹은 후행 공백과 빈 문자열 제거
#     refined_sentences = [sent.strip() for sent in refined_sentences if sent.strip()]
#     return refined_sentences

# def process_text_files(path):
#     all_files = glob.glob(os.path.join(path, '*.txt'))
#     filename_list = []
#     sent_list = []

#     for file_ in all_files:
#         with open(file_, 'r', encoding='utf-8') as f:
#             first_line = f.readline().strip()
#             remaining_text = f.read().strip()

#         sentences = custom_sentence_splitter(remaining_text)
#         sentences = [re.sub(r'([a-zA-Z])([\uAC00-\uD7A3])', r'\1 \2', sent) for sent in sentences]

#         for sent in sentences:
#             filename_list.append(first_line)
#             sent_list.append(sent)

#     return pd.DataFrame({'filename': filename_list, 'sent_text': sent_list})

# # '.' 기준 문장 추가 분리
# def split_sentences(text):
#     # Divide the text into sentences based on the period (.) following the Hangul characters
#     sentences = re.split('(?<=[\uAC00-\uD7A3])\.', text)
#     sentences = [sent.strip() for sent in sentences if sent]  # Select only non-space sentences and remove leading and trailing spaces
    
#     # Insert a space between English and Korean characters
#     sentences = [re.sub(r'([a-zA-Z])([\uAC00-\uD7A3])', r'\1 \2', sent) for sent in sentences]
#     return sentences

# # 괄호 안의 문자 처리
# def brackets_clean(text):
#             # 괄호 안의 숫자, 특수기호만 제거
#             clean1 = re.sub(r'\(([\d\W_]*?)\)', '()', text)
    
#             # 괄호와 문자간 띄어쓰기
#             clean2 = re.sub(r'([^\s])(\()', r'\1 \2', clean1)
#             clean3 = re.sub(r'(\))([^\s])', r'\1 \2', clean2)
    
#             return clean3

# # 단어 변환 1 : word_map에서 단어 길이가 긴 순으로 먼저 변환 실시, re-tubing은 리튜빙으로, tubing은 튜빙으로 인식되도록 함
# def replace(match):
#     return word_map[match.group(0)]

# def apply_replacement1(text):
#     # word_map의 키를 길이에 따라 내림차순으로 정렬합니다.
#     sorted_keys = sorted(word_map.keys(), key=len, reverse=True)
#     # lookbehind와 lookahead를 사용하여 단어의 일부만 매치되도록 패턴을 수정합니다.
#     pattern = re.compile('|'.join('(?<!\w){}(?!\w)'.format(re.escape(k)) for k in sorted_keys),re.IGNORECASE)
#     return pattern.sub(replace, text)

# # Function to apply the replacement within the text
# def apply_replacement2(text):
#     # Pattern that matches the words to be replaced even if they are part of a larger word
#     pattern = re.compile('|'.join(map(re.escape, word_map.keys())))
#     return pattern.sub(replace, text)

# # 부식율 인식 처리 함수
# def replace_with_words(text, word_map):
#     # Define a regular expression pattern for the intended formats
#     pattern = r'(\d+\.\d{2})\s*mm(?:/yr|/year)?|(\d+\.\d{2})\*mm(?:/yr|/year)?'

#     def replace(match):
#         # Extract the number and round it
#         num = round(float(match.group(1)), 2)
#         # Replace with corresponding word if exists
#         return f'{word_map.get(num, num)} mm/yr'  # Default to original number if not in word map

#     # Replace all occurrences in the text
#     return re.sub(pattern, replace, text)

# # 한글 추출 함수
# def extract_korean(text):
#     hangul = re.compile('[^ ㄱ-ㅣ 가-힣]')  
#     result = hangul.sub('', text)
    
#     return result

# # 머신러닝 적용 위한 토큰화 함수
# def tokenize(doc):
#     # pos 메서드를 사용하여 토큰화 및 품사 태깅, 정규화 및 기본형 변환 수행
#     return [word for word, tag in twitter.pos(doc, norm=True, stem=True)]

# # 딥러닝 적용 위한 토큰화 함수
# def tokenize1(doc):
#     # pos 메서드를 사용하여 토큰화 및 품사 태깅, 정규화 및 기본형 변환 수행
#     return [word for word, tag in twitter.pos(doc, norm=True, stem=False)]

# # 사전 토큰화된 데이터에 대한 생성기 함수를 정의    
# def tokens_generator(data):
#     for tokens in data:
#         yield tokens
        
# # 벡터화, 패딩        
# def sent2seq(token_list, vocab):
#     # 주어진 토큰 리스트를 사용하여 단어 사전에 기반한 인덱스의 리스트로 변환합니다.
#     # 각 토큰은 단어 사전에 해당하는 숫자 인덱스로 변환됩니다.
#     seq = [vocab[token] for token in token_list]
#     return seq

# def collate_function(batch, vocab):
#     label_list = []
#     sentence_list = []
#     # first = True
    
#     for (token_list, label) in batch:
#         # 토큰 리스트를 단어 사전의 인덱스로 변환한 뒤 텐서로 변환합니다.
#         seq = torch.tensor(np.array(sent2seq(token_list, vocab)))
#         sentence_list.append(seq)
#         label_list.append(label)
    
#     # pad_sequence를 사용하여 모든 시퀀스를 동일한 길이로 패딩합니다.
#     # 'batch_first=True'는 배치 크기가 반환된 텐서의 첫 번째 차원이 됨을 의미합니다.
#     # '<pad>' 토큰에 해당하는 인덱스를 사용하여 패딩합니다.
#     seq_list = pad_sequence(sentence_list, padding_value=vocab['<pad>'], batch_first=True)
    
#     # 레이블 리스트를 텐서로 변환합니다.
#     label_list = torch.tensor(label_list)
    
#     return torch.flip(seq_list, (1,)).to(torch.long), label_list
